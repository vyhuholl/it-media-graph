"""Every post reposted by more than two families, as an xlsx to read.

Exploratory tooling: not part of the package, not spec'd, not tested. Run
it, read the sheet, change the constants, run it again.

    uv sync --group data
    uv run notebooks/cited_posts.py

Ranked by how many *families* of affiliated channels reposted the post,
not by how many reposts it got. A network run by one author reposts its
own posts across every channel it owns, and counting those separately
would rank the largest family's routine self-distribution above a post
that travelled. One family, one voice — `sources` keeps the raw channel
count alongside, so a post carried by five channels of one family is
visibly that.

Reposts from inside the post's own family are dropped entirely, as in
``export_graph.py``: an author's channels quoting each other say nothing
about who reaches whom.

Forwards only. A mention names a channel, never a post — ``dst_msg_id``
is empty for it — so mentions cannot enter a post-level ranking at all.

An album arrives as one message id per part, so a forwarded album fills
several near-identical rows sharing a publication date, one per photo.
They are left as they are: which of the parts carried the text is
sometimes the interesting half, and collapsing them here would be a
counting decision the raw layer cannot support — ``grouped_id`` on an
edge belongs to the *referencing* message, not the referenced one.

Unlike ``export_graph.py`` this sheet is **not** filtered to seeds, and
it cannot be: a post is most interesting exactly when channels outside
the inventory's collected core keep reposting it. Non-seed channels have
no collected history, so their rows carry the edge-derived columns —
family count, sources, publication date — and nothing else. Blank text is
missing history, not a post without text. The price is that the sheet
names candidate and rejected channels together with their review status,
which makes it a local working file: it stays in `data/`, out of git and
out of anything published.

A threshold, not a top-N. ``families`` is a small integer: a handful of
posts reach the high single digits, and from two families downwards the
count stops separating anything — thousands of posts have been reposted
once, hundreds twice, and any fixed row count would cut somewhere inside
that band and hand the reader an arbitrary slice of it as if it were a
ranking. So the sheet takes every post above the band and nothing from
inside it, and its length is whatever the data says it is. Moving
``MIN_FAMILIES`` to 1 pulls the whole second band in at once; that is the
next honest cut, not a smaller one.

Age is the bias to read the order against. A post from last year has had
a year to be reposted, one from Monday has had days, and how far back the
edges reach at all is the backfill's cutoff rather than anything decided
here. This is "most reposted within the collected window", never "best".

The per-post metrics are one snapshot, read when the backfill walked the
channel: a year-old post is done growing, a three-day-old one is not.
They are here as context for a post you are already reading, not as a
column to sort the sheet by.
"""

from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import psycopg
from _db import dsn
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

DSN = dsn()
OUT = Path(__file__).resolve().parent.parent / "data" / "cited_posts.xlsx"
SHEET = "posts"

# Strictly more than this many families, so the value below is the
# highest one *excluded*. Raise it to shorten the sheet, lower it to
# admit the next band whole — never trade it for a row limit, which is
# the same threshold decided by whatever the sort did with the ties.
MIN_FAMILIES = 2

# The ranking itself. `channel_families` gives the family of each
# endpoint; a channel missing from the view is a family of one, hence the
# COALESCE — the same expression `export_graph.py` and the scorecard use.
TOP = """
    WITH refs AS (
        SELECT
            e.dst_channel_id AS channel_id,
            e.dst_msg_id AS msg_id,
            e.src_channel_id,
            e.dst_published_at,
            COALESCE(sf.family_key, e.src_channel_id) AS src_family,
            COALESCE(df.family_key, e.dst_channel_id) AS dst_family
        FROM edges e
        LEFT JOIN channel_families sf ON sf.channel_id = e.src_channel_id
        LEFT JOIN channel_families df ON df.channel_id = e.dst_channel_id
        WHERE e.kind = 'forward' AND e.dst_msg_id IS NOT NULL
    ),
    cited AS (
        SELECT
            channel_id,
            msg_id,
            COUNT(DISTINCT src_family) AS families,
            COUNT(DISTINCT src_channel_id) AS sources,
            MIN(dst_published_at) AS published
        FROM refs
        WHERE src_family <> dst_family
        GROUP BY channel_id, msg_id
    )
    SELECT
        channel_id,
        msg_id,
        families,
        sources,
        published
    FROM cited
    WHERE families > %(min_families)s
    ORDER BY families DESC, sources DESC, published DESC NULLS LAST,
             channel_id, msg_id
"""

# Reposts that never entered the ranking, for the report line.
DROPPED = """
    SELECT COUNT(*)
    FROM edges e
    LEFT JOIN channel_families sf ON sf.channel_id = e.src_channel_id
    LEFT JOIN channel_families df ON df.channel_id = e.dst_channel_id
    WHERE e.kind = 'forward' AND e.dst_msg_id IS NOT NULL
      AND COALESCE(sf.family_key, e.src_channel_id)
          = COALESCE(df.family_key, e.dst_channel_id)
"""

CHANNELS = """
    SELECT tg_id, title, username, status::text, kind::text
    FROM channels
    WHERE tg_id = ANY(%(ids)s)
"""

# The posts themselves, joined on the pairs the ranking returned:
# message ids are per-channel, so neither half selects a row alone and
# `unnest` of two arrays is what carries the pairs into SQL.
#
# Only channels with collected history have a row here at all, which is
# why this is a LEFT-join-shaped lookup in Python rather than a join in
# the query above.
#
# Reactions are the sum over the reaction types on a post. The
# `jsonb_typeof` guard is not defensive noise — the payload stores an
# absent field as JSON `null`, and `jsonb_array_elements` raises on a
# scalar rather than returning nothing.
POSTS = """
    SELECT
        r.channel_id,
        r.msg_id,
        (r.payload->>'date')::timestamptz,
        (r.payload->>'views')::bigint,
        CASE
            WHEN jsonb_typeof(r.payload->'reactions'->'results') = 'array'
            THEN COALESCE((
                SELECT SUM((one->>'count')::bigint)
                FROM jsonb_array_elements(
                    r.payload->'reactions'->'results'
                ) AS one
            ), 0)
            ELSE 0
        END,
        (r.payload->'replies'->>'replies')::bigint,
        (r.payload->>'forwards')::bigint,
        r.payload->>'message'
    FROM raw_messages r
    JOIN unnest(%(channels)s::bigint[], %(messages)s::bigint[])
         AS pair(channel_id, msg_id)
      ON pair.channel_id = r.channel_id AND pair.msg_id = r.msg_id
    WHERE r.payload->>'_' = 'Message'
"""

COLUMNS = [
    "families",
    "sources",
    "title",
    "username",
    "status",
    "kind",
    "msg_id",
    "published",
    "link",
    "views",
    "reactions",
    "comments",
    "forwards",
    "text",
]

SNAPSHOT_NOTE = (
    "SNAPSHOT, not a final count: read once, when the backfill walked "
    "this channel, so an old post is done growing and a recent one is "
    "not. Blank for a channel with no collected history. Context for a "
    "post you are reading, not a column to rank the sheet by."
)

NOTES = {
    "families": (
        "How many families of affiliated channels reposted this post — "
        "the ranking. One author's whole network counts once, and "
        "reposts from the post's own family are excluded entirely. Only "
        "channels in the inventory are seen, so this is a floor."
    ),
    "sources": (
        "How many individual channels reposted it. Larger than families "
        "means several of the reposting channels share an author; equal "
        "means that many independent voices."
    ),
    "status": (
        "Where the channel stands in review. Only 'seed' channels have "
        "collected history, which is why every post-level column is "
        "blank for the others."
    ),
    "kind": (
        "What the channel is, from the review. Filled in for seeds; "
        "blank elsewhere means unreviewed, not uncategorizable."
    ),
    "published": (
        "When the post was published, UTC — taken from the forward "
        "header, so it is known even for channels whose history was "
        "never collected."
    ),
    "link": (
        "Public t.me link. Blank when the channel has no username: no "
        "public link to the post exists then."
    ),
    "views": SNAPSHOT_NOTE,
    "reactions": SNAPSHOT_NOTE,
    "comments": (
        "Replies in the discussion thread, a snapshot like the rest. "
        "Blank means the channel has comments switched off — or that its "
        "history was never collected."
    ),
    "forwards": (
        "Telegram's own forward counter for the post: every repost "
        "anywhere, including private chats and channels outside the "
        "inventory. The ceiling that `sources` is measured against — a "
        "snapshot, and blank without collected history."
    ),
    "text": (
        "The post as collected, whitespace collapsed so the cell reads "
        "on one line. Blank for a channel with no collected history, and "
        "also for a post that is only a photo or a video."
    ),
}

WIDTHS = {
    "title": 38,
    "username": 22,
    "status": 11,
    "kind": 14,
    "msg_id": 10,
    "published": 18,
    "link": 34,
    "text": 100,
}

INT_FORMAT = "#,##0"
FORMATS = {
    "families": INT_FORMAT,
    "sources": INT_FORMAT,
    "msg_id": "0",
    "published": "yyyy-mm-dd hh:mm",
    "views": INT_FORMAT,
    "reactions": INT_FORMAT,
    "comments": INT_FORMAT,
    "forwards": INT_FORMAT,
}


def naive(value: datetime | None) -> datetime | None:
    """UTC without the tzinfo — Excel has no notion of a timezone."""
    if value is None:
        return None
    return value.astimezone(UTC).replace(tzinfo=None)


def one_line(text: str | None) -> str | None:
    """Collapse a post's whitespace, so the cell reads without wrapping."""
    return " ".join(text.split()) if text else None


def main() -> None:
    with psycopg.connect(DSN) as conn:
        top = conn.execute(TOP, {"min_families": MIN_FAMILIES}).fetchall()
        dropped = conn.execute(DROPPED).fetchone()
        intra_family = dropped[0] if dropped else 0

        pairs = {
            "channels": [row[0] for row in top],
            "messages": [row[1] for row in top],
        }
        channels = {
            row[0]: row[1:]
            for row in conn.execute(CHANNELS, {"ids": pairs["channels"]})
        }
        posts = {
            (row[0], row[1]): row[2:] for row in conn.execute(POSTS, pairs)
        }

    rows: list[dict[str, Any]] = []
    for channel_id, msg_id, families, sources, published in top:
        title, username, status, kind = channels.get(
            channel_id, (None, None, None, None)
        )
        date, views, reactions, comments, forwards, text = posts.get(
            (channel_id, msg_id), (None, None, None, None, None, None)
        )

        rows.append(
            {
                "families": families,
                "sources": sources,
                "title": title,
                "username": username,
                "status": status,
                "kind": kind,
                "msg_id": msg_id,
                "published": naive(published or date),
                "link": (
                    f"https://t.me/{username}/{msg_id}" if username else None
                ),
                "views": views,
                "reactions": reactions,
                "comments": comments,
                "forwards": forwards,
                "text": one_line(text),
            }
        )

    frame = pd.DataFrame(rows, columns=COLUMNS)
    write(frame)

    collected = int(frame["text"].notna().sum())
    print(f"{len(frame)} posts over {MIN_FAMILIES} families -> {OUT}")
    print(f"{intra_family} intra-family reposts dropped")
    print(f"{collected} of them have collected text")

    # How the sheet is distributed over the family counts — read it
    # before moving the threshold, since one step down is a band, not a
    # few rows.
    counts = frame["families"].value_counts().sort_index(ascending=False)
    spread = ", ".join(f"{n} at {key}" for key, n in counts.items())
    print(f"families: {spread}")


def write(frame: pd.DataFrame) -> None:
    """The sheet: sorted by families, header frozen, notes attached.

    A sort order *is* baked in here, unlike the scorecard — this sheet is
    a ranking, and the row order is the answer rather than an opinion
    about which column matters.
    """
    OUT.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name=SHEET, index=False)
        sheet = writer.sheets[SHEET]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for index, column in enumerate(frame.columns, start=1):
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = WIDTHS.get(column, 12)

            note = NOTES.get(column)
            if note:
                header = sheet.cell(row=1, column=index)
                header.comment = Comment(note, "cited_posts")

            number_format = FORMATS.get(column)
            if number_format:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(
                        row=row, column=index
                    ).number_format = number_format


if __name__ == "__main__":
    main()
