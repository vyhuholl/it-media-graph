"""Posts that beat their own channel's normal, as an xlsx to read.

Exploratory tooling: not part of the package, not spec'd, not tested. Run
it, read the sheets, change the constants, run it again.

    uv sync --group data
    uv run notebooks/anomalous_posts.py

The question is what a channel's audience found interesting, and the data
fights it: every metric is one snapshot, read when the backfill walked
the channel, on posts of every age. A year-old post has its final view
count, a three-day-old one is still climbing. Sort 200k posts by views
and the sheet ranks post age. Everything below exists to take the age
out.

**Maturity.** Only posts at least four weeks old *at the moment they
were read* enter the scoring — `fetched_at - date >= 28 days`, per row,
because the backfill ran over a week and a single cutoff date would be
four weeks for one channel and five for another. Four weeks is where
Telegram view counts have essentially stopped moving. The price is the
last month of history, which is the right price: a fresh post is a
real-time question, answered by snapshots *over* time, and this sheet has
exactly one snapshot.

**Two normalizations, answering two different questions.** They are kept
apart on purpose and never added together.

*Magnitude* — `views_z`, `views_lift`. A post's views against its own
channel's median, on a log scale. This is "what landed harder than this
channel's usual", corrected for channel size. It is still the most
age-sensitive column here, because views are what keep accruing longest.

*Type* — the rate columns, and the more interesting half. A ratio is
almost age-free and almost size-free: numerator and denominator grow
together, so the post's age largely cancels.

- `forward_rate` — readers carried it elsewhere instead of scrolling
  past. Audience behaviour, not the graph: it counts every forward
  anywhere, including private chats and channels nobody collects.
- `reaction_rate` — approval.
- `comment_rate` — argument. Agreement does not get typed out;
  a comment rate well above a channel's own baseline is a fight or a
  raw nerve far more often than it is enthusiasm.

A post high on forwards and a post high on comments are interesting in
opposite ways, so there is no combined score and no combined sheet. Four
sheets, four phenomena, four rankings.

The caveat the rates do not fix: forwards accrue more slowly than views,
so even a mature post's forward rate reads slightly low. Maturity
filtering covers most of it; nothing here covers the rest.

**How "anomalous" is decided.** Per channel, per metric:

- the rate is shrunk toward the channel's own baseline with
  `PSEUDO_VIEWS` views of prior — a post with 200 views and 6 forwards
  is not evidence of a 3% forward rate, and without this the sheets fill
  up with small posts whose ratio is one lucky reader wide;
- `_lift` is that shrunk rate over the channel's baseline rate, and
  reads directly: 4.0 is four times what this channel normally gets;
- `_z` is the same thing in units of the channel's *own* spread, so a
  channel whose rates swing wildly needs a bigger excursion to place
  than a channel whose rates never move. Median and a robust spread
  (MAD, or the upper half-spread where MAD collapses), never mean and
  standard deviation — one viral post moves those, and the outlier is
  what is being measured.

Ranking is on `_z`, at a threshold rather than a top-N, as in
``cited_posts.py``: an arbitrary row count would cut somewhere inside a
band of ties and present the slice as a ranking.

**Three guards, each removing a way to be wrong rather than a way to be
uninteresting.**

- A channel needs `MIN_CHANNEL_POSTS` mature posts before any of its
  posts can be scored. A median over a handful of posts is not a
  baseline.
- A post needs `MIN_COUNT` of the thing to place on that thing's sheet.
  Five comments is not a discussion however far above baseline it sits.
- A channel is scored on a metric only if its *median* post has some of
  it. Where half the posts have zero reactions the channel has reactions
  switched off or ignored, the baseline is a rounding error, and
  everything divided by it explodes. This is what keeps a vacancy feed
  whose posts get one reaction each from owning the reactions sheet; it
  costs roughly half the inventory on comments, which is what the data
  is.

**Albums are collapsed to one post.** 28% of collected messages are
album parts, and Telegram attaches reactions and comments to exactly one
part while giving every part its own views and forwards. Left alone,
every album donates several zero-engagement rows that drag its channel's
baseline down, and a viral album fills the top of a sheet with copies of
itself. So parts are merged on `grouped_id`: views and forwards are the
maximum over parts, reactions and comments the maximum (only one part
carries them), the text is the longest part's — the caption — and the id
is the first part's, which is what a t.me link to the album wants.
``cited_posts.py`` deliberately does *not* do this, and the difference is
real: there `grouped_id` belongs to the referencing message and says
nothing about the album being referenced, while here the message is the
one being measured.

**Reposts from an affiliated channel are dropped**, as intra-family
edges are in ``export_graph.py``. A network run by one author reposts
itself across every channel it owns; scoring those copies measures the
family's internal distribution and prints the same post several times.
Reposts of *unaffiliated* channels stay: choosing to carry someone
else's post is editorial work, and how this channel's audience responded
to it is exactly the question.

`families` and `sources` are the graph's view of the same post, carried
over from ``cited_posts.py`` — how many affiliation families, and how
many individual channels, were *seen* reposting it, family-internal
reposts excluded. Only collected channels can be seen, so both are
floors. Read them against `forwards`: a high forward count with no
families is a post that travelled somewhere this graph cannot see.

Seed channels only, so there is no `status` column — nothing else has
collected history. The sheet still carries titles and text from the
operator's own inventory, so it stays in `data/`, out of git and out of
anything published.
"""

from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path

import numpy as np
import pandas as pd
import psycopg
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

DSN = "postgresql://itgraph:itgraph@localhost:5433/itgraph"
OUT = Path(__file__).resolve().parent.parent / "data" / "anomalous_posts.xlsx"

# How old a post must be *when it was read* to count as done growing.
# Four weeks is where Telegram view counts flatten out.
MATURE_DAYS = 28

# A channel needs this many mature posts before it has a baseline worth
# dividing by.
MIN_CHANNEL_POSTS = 30

# And a post needs this many forwards / reactions / comments before its
# rate is allowed on that metric's sheet. Guards against a ratio built
# out of three events.
MIN_COUNT = 10

# Prior strength for the rate shrinkage, in views. A post with this many
# views is pulled halfway to its channel's baseline rate; the median
# collected post has around 3k, so a typical post keeps three quarters
# of its own evidence and a 200-view post keeps a sixth of it.
PSEUDO_VIEWS = 1000.0

# The threshold a post clears to reach a sheet, in units of its
# channel's own spread. Lower it to widen the sheets — but read the
# per-sheet counts the run prints first, since the bands are uneven.
MIN_Z = 3.0

# At most this many posts per channel per sheet. Without it a single
# channel having one extraordinary week takes over a whole sheet.
MAX_PER_CHANNEL = 5

# Which quantile stands in for one standard deviation when MAD
# collapses — as it does on a channel whose posts nearly all sit at the
# same value. 0.84 is where a normal distribution's +1σ falls.
UPPER_QUANTILE = 0.84

# Scaling that makes the MAD of a normal distribution equal its σ.
MAD_TO_SIGMA = 1.4826

# One row per collected message of a seed channel, mature at read time.
#
# `MessageService` rows are excluded: a "channel photo changed" event is
# not a post and carries no view count.
#
# Reactions are the sum over the reaction types on a post. The
# `jsonb_typeof` guard is not defensive noise — the payload stores an
# absent field as JSON `null`, and `jsonb_array_elements` raises on a
# scalar rather than returning nothing. NULL here means the channel
# published no reactions object at all, which is not the same as zero
# and is resolved per channel below.
#
# The album key is built in SQL and never in pandas: `grouped_id` is a
# full-width int64 and a column that has to hold NULLs would arrive as
# float64, where two distinct albums can round to one number.
POSTS = """
    SELECT
        r.channel_id,
        r.msg_id,
        COALESCE('g' || (r.payload->>'grouped_id'), 'm' || r.msg_id)
            AS album,
        c.title,
        c.username,
        c.kind::text,
        (r.payload->>'date')::timestamptz AS published,
        (r.payload->>'views')::float8 AS views,
        (r.payload->>'forwards')::float8 AS forwards,
        CASE
            WHEN jsonb_typeof(r.payload->'reactions'->'results') = 'array'
            THEN COALESCE((
                SELECT SUM((one->>'count')::bigint)
                FROM jsonb_array_elements(
                    r.payload->'reactions'->'results'
                ) AS one
            ), 0)::float8
        END AS reactions,
        (r.payload->'replies'->>'replies')::float8 AS comments,
        (r.payload->'fwd_from'->'from_id'->>'channel_id')::bigint
            AS reposted_from,
        r.payload->>'message' AS text
    FROM raw_messages r
    JOIN channels c ON c.tg_id = r.channel_id AND c.status = 'seed'
    WHERE r.payload->>'_' = 'Message'
      AND r.fetched_at - (r.payload->>'date')::timestamptz
          >= make_interval(days => %(mature_days)s)
"""

# Connected components of the confirmed affiliation pairs. A channel
# missing from the view is a family of one, hence the fallback below.
FAMILIES = """
    SELECT channel_id, family_key
    FROM channel_families
"""

# Who was observed reposting a seed post. Forwards only — a mention
# names a channel and never a post, so `dst_msg_id` is empty for it.
CITED = """
    SELECT e.dst_channel_id, e.dst_msg_id, e.src_channel_id
    FROM edges e
    JOIN channels d ON d.tg_id = e.dst_channel_id AND d.status = 'seed'
    WHERE e.kind = 'forward' AND e.dst_msg_id IS NOT NULL
"""

PART_COLUMNS = [
    "channel_id",
    "msg_id",
    "album",
    "title",
    "username",
    "kind",
    "published",
    "views",
    "forwards",
    "reactions",
    "comments",
    "reposted_from",
    "text",
]

# (payload field, column prefix) for the three rates.
RATES = [
    ("forwards", "forward"),
    ("reactions", "reaction"),
    ("comments", "comment"),
]

# One sheet per phenomenon, each ranked by its own z-score. Deliberately
# not one sheet with a combined score: a post that gets forwarded and a
# post that gets argued about are not two degrees of the same thing.
SHEETS = {
    "views": "views_z",
    "forwards": "forward_z",
    "reactions": "reaction_z",
    "comments": "comment_z",
}

COLUMNS = [
    "title",
    "username",
    "kind",
    "msg_id",
    "published",
    "link",
    "parts",
    "channel_posts",
    "views",
    "forwards",
    "reactions",
    "comments",
    "views_lift",
    "views_z",
    "forward_rate",
    "forward_lift",
    "forward_z",
    "reaction_rate",
    "reaction_lift",
    "reaction_z",
    "comment_rate",
    "comment_lift",
    "comment_z",
    "families",
    "sources",
    "text",
]

LIFT_NOTE = (
    "How many times this channel's own baseline rate the post reached, "
    "after shrinking the ratio toward that baseline with "
    f"{PSEUDO_VIEWS:.0f} views of prior — so a post with few views has "
    "to be extreme, not lucky, to score. Reads directly: 4.0 is four "
    "times normal for this channel."
)

Z_NOTE = (
    "The same excess, in units of how much this channel's rate normally "
    "varies (median and a robust spread, never mean and σ). A channel "
    "whose posts swing needs a bigger excursion to place here than one "
    "whose rate never moves. This column is the sheet's ranking."
)

BLANK_NOTE = (
    "Blank means not scored, never zero: the channel's median post has "
    f"none of this at all, or the post has under {MIN_COUNT} of them, "
    f"or the channel has under {MIN_CHANNEL_POSTS} mature posts."
)

NOTES = {
    "kind": "What the channel is, from the review.",
    "msg_id": (
        "The post, in Telegram's per-channel numbering. For an album, "
        "the first part — which is what a link to the album wants."
    ),
    "published": "When the post was published, UTC.",
    "parts": (
        "How many messages the post is. Above 1 it is an album, merged "
        "into this one row: views and forwards are the largest over the "
        "parts, reactions and comments come from the single part "
        "Telegram attaches them to, and the text is the caption."
    ),
    "channel_posts": (
        "Mature posts this channel's baselines were computed over, "
        "albums counted once. How solid every normalized column below "
        f"is; under {MIN_CHANNEL_POSTS} the channel is not scored at all."
    ),
    "views": (
        "SNAPSHOT, read once when the backfill walked this channel. Only "
        f"posts already {MATURE_DAYS} days old when read are here, so it "
        "is close to final — but views are the slowest metric to settle, "
        "and this is the column where age still shows."
    ),
    "forwards": (
        "Telegram's own forward counter: every repost anywhere, "
        "including private chats and channels outside the inventory. A "
        "snapshot, like the rest."
    ),
    "reactions": (
        "Reactions summed over the emoji, a snapshot. Blank means the "
        "channel publishes no reactions at all — not that nobody reacted."
    ),
    "comments": (
        "Replies in the discussion thread, a snapshot. Blank means the "
        "channel has comments switched off."
    ),
    "views_lift": (
        "Views over this channel's median post's views. 10.0 is ten "
        "times the channel's normal reach. The most age-sensitive "
        "column in the sheet — an older post has had longer to collect."
    ),
    "views_z": (
        "How far above the channel's median this post's views sit, in "
        "units of the channel's own spread, on a log scale. Ranks the "
        "'views' sheet. Read it against views_lift: a huge z with a "
        "small lift is a channel whose posts are usually identical."
    ),
    "forward_rate": (
        "forwards / views for this post: the share of readers who "
        "carried it somewhere else instead of scrolling past. Nearly "
        "age-free, since both halves grow together. " + BLANK_NOTE
    ),
    "forward_lift": LIFT_NOTE,
    "forward_z": Z_NOTE + " Ranks the 'forwards' sheet.",
    "reaction_rate": (
        "reactions / views: the share of readers who approved out loud. "
        + BLANK_NOTE
    ),
    "reaction_lift": LIFT_NOTE,
    "reaction_z": Z_NOTE + " Ranks the 'reactions' sheet.",
    "comment_rate": (
        "comments / views: the share of readers who typed something. "
        "Agreement does not get typed — a rate well above the channel's "
        "own baseline is an argument or a raw nerve more often than it "
        "is enthusiasm. Can exceed 1.0: the discussion group has its own "
        "audience. " + BLANK_NOTE
    ),
    "comment_lift": LIFT_NOTE,
    "comment_z": Z_NOTE + " Ranks the 'comments' sheet.",
    "families": (
        "How many families of affiliated channels were seen reposting "
        "this post, reposts from its own family excluded — the same "
        "count `cited_posts.py` ranks by. Only channels in the "
        "inventory are visible, so it is a floor. Blank means none was "
        "seen, which is the usual case. Read against `forwards`: many "
        "forwards and no families is a post that travelled where this "
        "graph cannot see."
    ),
    "sources": (
        "How many individual channels reposted it. Larger than families "
        "means several of them share an author."
    ),
    "text": (
        "The post as collected, whitespace collapsed. Blank for a post "
        "that is only a photo or a video."
    ),
}

WIDTHS = {
    "title": 34,
    "username": 20,
    "kind": 12,
    "msg_id": 9,
    "published": 17,
    "link": 32,
    "parts": 7,
    "text": 100,
}

INT_FORMAT = "#,##0"
FORMATS = {
    "msg_id": "0",
    "published": "yyyy-mm-dd hh:mm",
    "parts": "0",
    "channel_posts": INT_FORMAT,
    "views": INT_FORMAT,
    "forwards": INT_FORMAT,
    "reactions": INT_FORMAT,
    "comments": INT_FORMAT,
    "views_lift": "0.0",
    "views_z": "0.0",
    "forward_rate": "0.0000",
    "forward_lift": "0.0",
    "forward_z": "0.0",
    "reaction_rate": "0.0000",
    "reaction_lift": "0.0",
    "reaction_z": "0.0",
    "comment_rate": "0.0000",
    "comment_lift": "0.0",
    "comment_z": "0.0",
    "families": INT_FORMAT,
    "sources": INT_FORMAT,
}


def naive(value: datetime | None) -> datetime | None:
    """UTC without the tzinfo — Excel has no notion of a timezone."""
    if value is None or pd.isna(value):
        return None
    return value.astimezone(UTC).replace(tzinfo=None)


def one_line(text: str | None) -> str | None:
    """Collapse a post's whitespace, so the cell reads without wrapping."""
    return " ".join(text.split()) if text else None


def spread(
    values: pd.Series, groups: pd.Series
) -> tuple[pd.Series, pd.Series]:
    """Median and a robust spread of ``values`` within each group.

    MAD first, because it is the estimator that a single viral post
    cannot move — which matters when the viral post is the thing being
    measured. It collapses to zero on a channel whose posts nearly all
    carry the same value, so the upper half-spread (the 84th percentile
    over the median, where a normal distribution's +1σ sits) stands in
    whenever it is the larger of the two. A spread that is still zero
    leaves the z-score blank rather than infinite: that channel has no
    variation to measure an outlier against.
    """
    median = values.groupby(groups).transform("median")
    deviation = (values - median).abs()
    mad = deviation.groupby(groups).transform("median") * MAD_TO_SIGMA
    upper = values.groupby(groups).transform(
        lambda group: group.quantile(UPPER_QUANTILE)
    )
    widest = pd.concat([mad, upper - median], axis=1).max(axis=1)
    return median, widest.where(widest > 0)


def collapse_albums(parts: pd.DataFrame) -> pd.DataFrame:
    """One row per post, album parts merged — see the module docstring.

    Sorted by text length first so that plain ``last`` picks the album's
    caption; every other aggregation here is order-independent.
    """
    ordered = parts.assign(
        length=parts["text"].fillna("").str.len()
    ).sort_values("length", kind="stable")

    return (
        ordered.groupby(["channel_id", "album"], sort=False)
        .agg(
            msg_id=("msg_id", "min"),
            title=("title", "last"),
            username=("username", "last"),
            kind=("kind", "last"),
            published=("published", "min"),
            views=("views", "max"),
            forwards=("forwards", "max"),
            reactions=("reactions", "max"),
            comments=("comments", "max"),
            text=("text", "last"),
            parts=("msg_id", "size"),
        )
        .reset_index()
    )


def score(posts: pd.DataFrame) -> pd.DataFrame:
    """Add the lift and z columns, one pair per metric."""
    channel = posts["channel_id"]

    # Magnitude, on a log scale: view counts span three orders of
    # magnitude inside one channel and are roughly log-normal there,
    # while on the raw scale a median-and-MAD would call every large
    # post an outlier.
    log_views = np.log(posts["views"])
    median, widest = spread(log_views, channel)
    posts["views_lift"] = posts["views"] / np.exp(median)
    posts["views_z"] = (log_views - median) / widest

    for column, name in RATES:
        # The channel is only scored on this metric if its median post
        # has some of it; otherwise the baseline is a rounding error and
        # every ratio taken against it explodes.
        typical = posts.groupby("channel_id")[column].transform("median")
        counted = posts[column].where(typical > 0)

        # The baseline is pooled over the channel, not a median of
        # per-post ratios: it is the prior mean the shrinkage below
        # pulls toward, and a median of ratios is not that.
        total = counted.groupby(channel).transform("sum")
        seen = posts["views"].where(counted.notna())
        baseline = total / seen.groupby(channel).transform("sum")

        shrunk = (counted + PSEUDO_VIEWS * baseline) / (
            posts["views"] + PSEUDO_VIEWS
        )
        log_rate = np.log(shrunk)
        median, widest = spread(log_rate, channel)

        posts[f"{name}_rate"] = counted / posts["views"]
        posts[f"{name}_lift"] = shrunk / baseline
        posts[f"{name}_z"] = ((log_rate - median) / widest).where(
            counted >= MIN_COUNT
        )

    return posts.replace([np.inf, -np.inf], np.nan)


def main() -> None:
    with psycopg.connect(DSN) as conn:
        rows = conn.execute(POSTS, {"mature_days": MATURE_DAYS}).fetchall()
        families = {
            channel_id: key for channel_id, key in conn.execute(FAMILIES)
        }
        cited = conn.execute(CITED).fetchall()

    parts = pd.DataFrame(rows, columns=PART_COLUMNS)
    print(f"{len(parts)} mature messages from seed channels")

    # Which album each message id belongs to, built before anything is
    # dropped, so an edge pointing at a discarded part still resolves.
    album_of = {
        (channel_id, msg_id): album
        for channel_id, msg_id, album in zip(
            parts["channel_id"], parts["msg_id"], parts["album"], strict=True
        )
    }

    def family(channel_id: int) -> int:
        return families.get(channel_id, channel_id)

    own = parts["channel_id"].map(family)
    from_family = parts["reposted_from"].map(
        lambda source: None if pd.isna(source) else family(int(source))
    )
    affiliated = from_family.notna() & (from_family == own)
    parts = parts[~affiliated]
    print(f"{int(affiliated.sum())} reposts of affiliated channels dropped")

    posts = collapse_albums(parts)
    albums = int((posts["parts"] > 1).sum())
    print(f"{len(posts)} posts after merging {albums} albums")

    # A missing reactions or comments object means "this channel does
    # not do this" on a channel that never publishes one, and "nobody
    # did" on a channel that usually does. Only the second is a zero.
    for column in ("reactions", "comments"):
        offered = posts.groupby("channel_id")[column].transform("count") > 0
        posts[column] = posts[column].fillna(0.0).where(offered)

    # A post with no view count cannot be normalized by one, so it is
    # gone before the baselines are counted rather than after.
    posts = posts[posts["views"].notna() & (posts["views"] > 0)]
    posts["channel_posts"] = posts.groupby("channel_id")["views"].transform(
        "size"
    )
    thin = posts["channel_posts"] < MIN_CHANNEL_POSTS
    posts = posts[~thin].copy()
    print(
        f"{len(posts)} posts scored over "
        f"{posts['channel_id'].nunique()} channels "
        f"({int(thin.sum())} dropped under {MIN_CHANNEL_POSTS} posts)"
    )

    posts = score(posts)

    sources: defaultdict[tuple[int, str], set[int]] = defaultdict(set)
    seen_families: defaultdict[tuple[int, str], set[int]] = defaultdict(set)
    intra_family = 0
    for dst_channel, dst_msg, src_channel in cited:
        album = album_of.get((dst_channel, dst_msg))
        if album is None:
            continue
        if family(src_channel) == family(dst_channel):
            intra_family += 1
            continue
        sources[(dst_channel, album)].add(src_channel)
        seen_families[(dst_channel, album)].add(family(src_channel))
    print(f"{intra_family} intra-family reposts dropped from `families`")

    keys = list(zip(posts["channel_id"], posts["album"], strict=True))
    posts["families"] = [
        len(seen_families.get(key, ())) or None for key in keys
    ]
    posts["sources"] = [len(sources.get(key, ())) or None for key in keys]

    posts["published"] = posts["published"].map(naive)
    posts["text"] = posts["text"].map(one_line)
    posts["link"] = [
        f"https://t.me/{username}/{msg_id}" if username else None
        for username, msg_id in zip(
            posts["username"], posts["msg_id"], strict=True
        )
    ]

    sheets = {name: pick(posts, key) for name, key in SHEETS.items()}
    write(sheets)

    print(f"-> {OUT}")
    for name, frame in sheets.items():
        scored = int(posts[SHEETS[name]].notna().sum())
        over = int((posts[SHEETS[name]] >= MIN_Z).sum())
        print(
            f"  {name}: {len(frame)} rows over "
            f"{frame['username'].nunique()} channels "
            f"({over} above z {MIN_Z}, capped at {MAX_PER_CHANNEL} per "
            f"channel; {scored} posts were scorable at all)"
        )


def pick(posts: pd.DataFrame, key: str) -> pd.DataFrame:
    """One sheet: everything over the threshold, capped per channel.

    A threshold rather than a top-N, as in ``cited_posts.py``: a fixed
    row count cuts somewhere inside a band of near-identical scores and
    hands the reader an arbitrary slice of it as if it were a ranking.
    The per-channel cap is the one exception, and it is a cap on
    repetition rather than on rank — a channel that had one remarkable
    week would otherwise take the whole sheet.
    """
    ranked = posts[posts[key] >= MIN_Z].sort_values(key, ascending=False)
    capped = ranked.groupby("channel_id", sort=False).head(MAX_PER_CHANNEL)
    return capped[COLUMNS].reset_index(drop=True)


def write(sheets: dict[str, pd.DataFrame]) -> None:
    """The workbook: one sheet per phenomenon, header frozen, notes on.

    A sort order *is* baked in, unlike the scorecard — each sheet is a
    ranking, and the row order is the answer rather than an opinion
    about which column matters.
    """
    OUT.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
            sheet = writer.sheets[name]
            sheet.freeze_panes = "D2"
            sheet.auto_filter.ref = sheet.dimensions

            for index, column in enumerate(frame.columns, start=1):
                letter = get_column_letter(index)
                sheet.column_dimensions[letter].width = WIDTHS.get(column, 13)

                note = NOTES.get(column)
                if note:
                    header = sheet.cell(row=1, column=index)
                    header.comment = Comment(note, "anomalous_posts")

                number_format = FORMATS.get(column)
                if number_format:
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(
                            row=row, column=index
                        ).number_format = number_format


if __name__ == "__main__":
    main()
