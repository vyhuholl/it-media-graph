"""One row per seed channel, as an xlsx to open and sort by hand.

Exploratory tooling: not part of the package, not spec'd, not tested. Run
it, sort the sheet, change the constants, run it again.

    uv sync --group data
    uv run notebooks/channel_scorecard.py

Only seed channels get a row: nothing else has collected history, so
every per-post column would be empty and every rate a division by zero.
Non-seed channels still count as endpoints — a seed that reposts one is
reposting *someone*, and pretending otherwise would understate its
outgoing variety.

There is no time window: every column covers the whole collected
history, and how far back that reaches is the backfill's cutoff rather
than anything decided here. It can differ per channel, which is why
``posts`` is a collection depth as much as a publishing rate — the
medians and rates below are unaffected, being per-post.

Three things this script refuses to do, each for the same reason.

Post metrics are one snapshot, taken whenever the backfill happened to
walk the channel. A post from last July has its final view count; one
from Monday is still climbing. So: engagement is normalized inside the
channel (a post's reactions over that same post's views, both read at the
same instant, so the age cancels), the absolute medians are reported raw
but marked as snapshots in the header comments, and there is no
"top posts by views" here at all — a ranking of posts across channels
would be a ranking of post age. Channel-level aggregates only.

Medians, never means. Two orders of magnitude separate the largest seed
channel from the smallest, and one viral post moves a mean but not a
median.

Edges inside one family of affiliated channels are dropped, as in
``export_graph.py``. A network run by one author reposts itself
constantly; left in, that traffic makes the largest family look like the
most outward-facing channel in the inventory.
"""

import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
import psycopg
from _db import dsn
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

DSN = dsn()
OUT = (
    Path(__file__).resolve().parent.parent / "data" / "channel_scorecard.xlsx"
)
SHEET = "channels"

# Below this many collected posts, the rate columns are left blank rather
# than reported. A median over four posts is not an engagement rate, and
# a number on the screen gets sorted on whatever it means.
MIN_POSTS_FOR_RATES = 10

# Which references count as an outgoing edge. Both kinds, as in the graph
# export — a mention is a weaker signal than a repost but the same act.
EDGE_KINDS = ("forward", "mention")

CHANNELS = """
    SELECT
        c.tg_id,
        c.title,
        c.username,
        c.kind::text,
        (rc.payload->'full_chat'->>'participants_count')::bigint
    FROM channels c
    LEFT JOIN raw_channels rc ON rc.channel_id = c.tg_id
    WHERE c.status = 'seed'
"""

# Connected components of the confirmed affiliation pairs. A channel
# missing from the view is a family of one, hence the fallback below.
FAMILIES = """
    SELECT channel_id, family_key
    FROM channel_families
"""

# Per-post metrics, aggregated in Postgres because `percentile_cont` is
# already there and the raw layer is 200k rows.
#
# `MessageService` rows are excluded: a "channel photo changed" event is
# not a post and carries no view count.
#
# Reactions are the sum over the reaction types on a post. The
# `jsonb_typeof` guard is not defensive noise — the payload stores an
# absent field as JSON `null`, and `jsonb_array_elements` raises on a
# scalar rather than returning nothing.
#
# The rate columns are the median of a *per-post* ratio, not one median
# divided by another: pairing each post's reactions with its own views is
# what cancels the snapshot age, and a ratio of two independent medians
# would not.
POSTS = """
    WITH posts AS (
        SELECT
            r.channel_id,
            (r.payload->>'views')::bigint AS views,
            (r.payload->>'forwards')::bigint AS forwards,
            CASE
                WHEN jsonb_typeof(r.payload->'reactions'->'results')
                     = 'array'
                THEN COALESCE((
                    SELECT SUM((one->>'count')::bigint)
                    FROM jsonb_array_elements(
                        r.payload->'reactions'->'results'
                    ) AS one
                ), 0)
                ELSE 0
            END AS reactions,
            (r.payload->'replies'->>'replies')::bigint AS comments
        FROM raw_messages r
        JOIN channels c ON c.tg_id = r.channel_id AND c.status = 'seed'
        WHERE r.payload->>'_' = 'Message'
    )
    SELECT
        channel_id,
        COUNT(*) AS posts,
        percentile_cont(0.5) WITHIN GROUP (ORDER BY views),
        percentile_cont(0.5) WITHIN GROUP (ORDER BY reactions),
        percentile_cont(0.5) WITHIN GROUP (ORDER BY forwards),
        percentile_cont(0.5) WITHIN GROUP (
            ORDER BY reactions::float8 / views
        ) FILTER (WHERE views > 0),
        percentile_cont(0.5) WITHIN GROUP (
            ORDER BY forwards::float8 / views
        ) FILTER (WHERE views > 0 AND forwards IS NOT NULL),
        percentile_cont(0.5) WITHIN GROUP (
            ORDER BY comments::float8 / views
        ) FILTER (WHERE views > 0 AND comments IS NOT NULL)
    FROM posts
    GROUP BY channel_id
"""

EDGES = """
    SELECT src_channel_id, dst_channel_id
    FROM edges
    WHERE kind::text = ANY(%(kinds)s)
"""

# Which channels have stored history at all. A channel with none cannot
# be *seen* reposting back, so counting it as a target that failed to
# reciprocate would measure collection coverage, not reciprocity.
COLLECTED = """
    SELECT DISTINCT channel_id
    FROM raw_messages
"""

COLUMNS = [
    "title",
    "username",
    "kind",
    "family_key",
    "subscribers",
    "posts",
    "median_views",
    "median_reactions",
    "median_forwards",
    "reaction_rate",
    "forward_rate",
    "comment_rate",
    "out_variety",
    "out_targets",
    "in_sources",
    "reciprocity",
]

SNAPSHOT_NOTE = (
    "SNAPSHOT, not a final count. Every post was measured once, when the "
    "backfill walked this channel, so a year-old post is done growing and "
    "a three-day-old one is not. Comparable across channels only as an "
    "order of magnitude; for engagement use the rate columns, which "
    "divide by the same post's own views."
)

NOTES = {
    "family_key": (
        "The family of affiliated channels this one belongs to, as the "
        "smallest channel id in it. A label, not a main channel. A "
        "channel with no confirmed affiliation is its own family, so the "
        "key is its own id."
    ),
    "subscribers": (
        "participants_count from the last GetFullChannelRequest. Blank "
        "means `itgraph metadata` has not covered this channel yet, not "
        "that it has no subscribers."
    ),
    "posts": (
        "Posts in the collected history, service messages excluded. How "
        "deep that history goes is the backfill's cutoff and can differ "
        "per channel, so this is a collection depth as much as a "
        "publishing rate — compare two channels on it only after "
        "checking they were walked to the same depth."
    ),
    "median_views": SNAPSHOT_NOTE,
    "median_reactions": SNAPSHOT_NOTE,
    "median_forwards": SNAPSHOT_NOTE,
    "reaction_rate": (
        "Median over posts of (that post's reactions / that post's "
        "views). Both numbers come from one snapshot, so post age "
        "cancels and the result compares across channels of any size. "
        f"Blank under {MIN_POSTS_FOR_RATES} posts."
    ),
    "forward_rate": (
        "Median over posts of (that post's forwards / that post's "
        "views) — how often readers carry a post elsewhere. Blank under "
        f"{MIN_POSTS_FOR_RATES} posts."
    ),
    "comment_rate": (
        "Median over posts of (that post's comments / that post's "
        "views). Counted only over posts that have a comment thread, so "
        "a channel with comments switched off is blank rather than zero."
    ),
    "out_variety": (
        "Shannon entropy (bits) of how this channel's outgoing "
        "references are spread over their targets: 0 means every "
        "reference goes to one channel, higher means spread evenly over "
        "more. Reads against out_targets — 3 bits over 8 targets is an "
        "even spread, 3 bits over 200 is a channel with a favourite. "
        "Blank when there are no outgoing references."
    ),
    "out_targets": (
        "Distinct channels this one references. References inside its "
        "own family are excluded."
    ),
    "in_sources": (
        "Distinct channels that reference this one. Only channels with "
        "collected history can be counted as a source, so this is a "
        "floor, not a total. Family excluded."
    ),
    "reciprocity": (
        "Share of this channel's targets that reference it back. The "
        "denominator counts only targets whose own history was "
        "collected — a channel nobody has walked cannot be observed "
        "reciprocating. Blank when no target qualifies."
    ),
}

WIDTHS = {"title": 38, "username": 22, "kind": 12, "family_key": 13}

INT_FORMAT = "#,##0"
FORMATS = {
    "family_key": "0",
    "subscribers": INT_FORMAT,
    "posts": INT_FORMAT,
    "median_views": INT_FORMAT,
    "median_reactions": INT_FORMAT,
    "median_forwards": INT_FORMAT,
    "reaction_rate": "0.0000",
    "forward_rate": "0.0000",
    "comment_rate": "0.0000",
    "out_variety": "0.00",
    "out_targets": INT_FORMAT,
    "in_sources": INT_FORMAT,
    "reciprocity": "0.000",
}


def entropy(targets: Counter[int]) -> float | None:
    """Shannon entropy, in bits, of a distribution over targets."""
    total = sum(targets.values())
    if not total:
        return None
    return -sum(
        (n / total) * math.log2(n / total) for n in targets.values() if n
    )


def main() -> None:
    params = {"kinds": list(EDGE_KINDS)}

    outgoing: dict[int, Counter[int]] = defaultdict(Counter)
    incoming: dict[int, set[int]] = defaultdict(set)
    intra_family = 0

    with psycopg.connect(DSN) as conn:
        families = {
            channel_id: key for channel_id, key in conn.execute(FAMILIES)
        }
        collected = {row[0] for row in conn.execute(COLLECTED)}
        channels = conn.execute(CHANNELS).fetchall()
        stats = {row[0]: row[1:] for row in conn.execute(POSTS)}

        for src, dst in conn.execute(EDGES, params):
            if families.get(src, src) == families.get(dst, dst):
                intra_family += 1
                continue
            outgoing[src][dst] += 1
            incoming[dst].add(src)

    rows: list[dict[str, Any]] = []
    for tg_id, title, username, kind, subscribers in channels:
        posts, views, reactions, forwards, r_rate, f_rate, c_rate = stats.get(
            tg_id, (0, None, None, None, None, None, None)
        )
        thin = posts < MIN_POSTS_FOR_RATES
        targets = outgoing.get(tg_id, Counter())

        # Only targets that were walked can be observed reposting back.
        answerable = [dst for dst in targets if dst in collected]
        back = sum(1 for dst in answerable if tg_id in outgoing.get(dst, ()))

        rows.append(
            {
                "title": title,
                "username": username,
                "kind": kind,
                "family_key": families.get(tg_id, tg_id),
                "subscribers": subscribers,
                "posts": posts,
                "median_views": views,
                "median_reactions": reactions,
                "median_forwards": forwards,
                "reaction_rate": None if thin else r_rate,
                "forward_rate": None if thin else f_rate,
                "comment_rate": None if thin else c_rate,
                "out_variety": entropy(targets),
                "out_targets": len(targets),
                "in_sources": len(incoming.get(tg_id, ())),
                "reciprocity": (
                    back / len(answerable) if answerable else None
                ),
            }
        )

    frame = pd.DataFrame(rows, columns=COLUMNS)
    write(frame)

    without_metadata = int(frame["subscribers"].isna().sum())
    thin = int((frame["posts"] < MIN_POSTS_FOR_RATES).sum())
    print(f"{len(frame)} seed channels -> {OUT}")
    print(f"{intra_family} intra-family edges dropped")
    print(f"{without_metadata} channels without a metadata payload")
    print(f"{thin} channels under {MIN_POSTS_FOR_RATES} posts, rates blank")


def write(frame: pd.DataFrame) -> None:
    """The sheet, unsorted: header frozen, autofilter on, notes attached.

    No sort order is baked in — the point of an xlsx is that the reader
    picks the column and the direction, and any default here would be an
    opinion about what matters.
    """
    OUT.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name=SHEET, index=False)
        sheet = writer.sheets[SHEET]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for index, column in enumerate(frame.columns, start=1):
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = WIDTHS.get(column, 15)

            note = NOTES.get(column)
            if note:
                header = sheet.cell(row=1, column=index)
                header.comment = Comment(note, "channel_scorecard")

            number_format = FORMATS.get(column)
            if number_format:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(
                        row=row, column=index
                    ).number_format = number_format


if __name__ == "__main__":
    main()
