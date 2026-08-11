"""Who sits with whom: Leiden over the seed reference graph, as xlsx.

Exploratory tooling: not part of the package, not spec'd, not tested. Run
it, read the sheet, change the constants, run it again.

    uv sync --group data
    uv run notebooks/clusters.py

Writes three sheets and a GEXF for Gephi. ``notebooks/cluster_map.py``
imports the clustering from here rather than repeating it.

**Both edge kinds, not forwards alone.** A repost graph is the obvious
thing to cluster and it is too thin to do it: only 279 of 544 seed
channels repost another seed at all, so half the inventory would have no
cluster for lack of an edge rather than for lack of a crowd. Adding
mentions brings the graph to 470 connected channels. A mention is a
weaker claim than a repost — "I am talking about this" against "I am
republishing this" — but it is the same act of pointing at someone, and
it is the difference between clustering the inventory and clustering half
of it.

**The two kinds are balanced before they are added.** There are 3229
units of decayed mention weight against 2044 of forward, so a plain sum
would let mentions decide the partition by being more numerous. Each kind
is scaled to the same total mass first, which is the only defensible
reading of "both count".

**Pair weight is square-rooted.** Half the connected pairs in this graph
were observed exactly once and the busiest was observed 154 times.
Undamped, that one pair would weigh as much as 154 separate
relationships, and the clusters would form around individual friendships
rather than crowds.

**Edges inside one family of affiliated channels are dropped**, as in
every other script here — a third of all seed-to-seed traffic is one
author reposting themselves, and it says nothing about who influences
whom.

Time decay is present and barely matters: half-lives of 90, 180 and 365
days, and no decay at all, produce partitions agreeing at NMI 0.76-0.81.
It is kept because a two-year-old repost genuinely is weaker evidence
than a recent one, not because the answer turns on it.

**Resolution is what the answer turns on**, and it is a choice rather
than a measurement: 12 clusters at 0.8 against 31 at 3.0. The default
below is where the graph stops having one 106-channel blob in the middle
of it — at 1.6 the largest cluster is 12% of the nodes and 15 clusters
hold at least five channels. Move it and re-read; there is no correct
value, only a granularity worth looking at.

**Chats are excluded.** History is not collected from them, so they can
have incoming references and never an outgoing one — a node placed
entirely by who talks about it.

Which is also the sheet's main blind spot, and it is worth stating in
full: a channel is placed by the references *observed*, and references
are only observed from channels that were walked. A cluster is therefore
a claim about the collected core, and a channel with few collected
neighbours sits where its handful of edges put it. That is what
``stability`` is for — read it before believing a label.
"""

import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import UTC, datetime
from functools import cache
from pathlib import Path
from typing import Any

import igraph as ig
import leidenalg
import networkx as nx
import pandas as pd
import psycopg
import spacy
from _db import dsn
from channel_style import style_frame
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from spacy.lang.en.stop_words import STOP_WORDS as EN_STOP_WORDS
from spacy.lang.ru.stop_words import STOP_WORDS as RU_STOP_WORDS
from spacy.language import Language

DSN = dsn()
DATA = Path(__file__).resolve().parent.parent / "data"
OUT = DATA / "clusters.xlsx"
GEXF = DATA / "clusters.gexf"

HALF_LIFE_DAYS = 180.0

# See the module docstring: the granularity worth looking at, not a
# measured constant. Raise it for more, smaller crowds.
RESOLUTION = 1.6

# How many times to re-run Leiden with a different seed to find out
# whether a channel's placement is a fact about the graph or about the
# random order the algorithm happened to visit it in. Fifty runs over 470
# nodes costs a couple of seconds.
STABILITY_RUNS = 50

# A cluster smaller than this is not a crowd, and its members are left
# unlabelled rather than given a label that describes three channels.
MIN_CLUSTER = 5

# And a channel joined to the graph by a single edge is placed entirely
# by that edge. Blank is honest; a label is not.
MIN_PARTNERS = 2

# How many distinctive terms name a cluster, and how many times a term
# must appear in it before it is allowed to be one of them.
NAME_TERMS = 5
MIN_TERM_COUNT = 3

# Bridges: channels worth reading as connectors rather than members. A
# low inside_share on three edges is noise, so the sheet asks for a
# channel with enough links for the share to mean something.
BRIDGE_MIN_PARTNERS = 5

NODES = """
    SELECT
        c.tg_id,
        c.title,
        c.username,
        c.kind::text,
        rc.payload->'full_chat'->>'about'
    FROM channels c
    LEFT JOIN raw_channels rc ON rc.channel_id = c.tg_id
    WHERE c.status = 'seed' AND NOT c.is_chat
"""

EDGES = """
    SELECT e.src_channel_id, e.dst_channel_id, e.kind::text, e.published_at
    FROM edges e
    JOIN channels s ON s.tg_id = e.src_channel_id
                   AND s.status = 'seed' AND NOT s.is_chat
    JOIN channels d ON d.tg_id = e.dst_channel_id
                   AND d.status = 'seed' AND NOT d.is_chat
    WHERE e.src_channel_id <> e.dst_channel_id
"""

FAMILIES = """
    SELECT channel_id, family_key
    FROM channel_families
"""

KINDS = ("forward", "mention")

# Function words and channel boilerplate, which are frequent everywhere
# and distinctive nowhere. spaCy's two standard lists carry the function
# words; the set below carries what is boilerplate only in this corpus —
# a Telegram channel description saying it is a Telegram channel.
#
# Imported from `spacy.lang`, not read off a loaded model. Both are the
# same frozenset, but `spacy.load` builds the entire pipeline to hand
# back a static attribute that needs none of it — the models are loaded
# separately, by `lemmatize`, which is the part that actually needs one.
#
# The inflected Russian forms below are now mostly redundant, since terms
# are counted as lemmas and `канал` catches `канала` and `каналу` on its
# own. They are kept because the filter checks the surface form too: a
# token the lemmatizer mis-tags still gets caught by its spelling.
STOPWORDS = (
    EN_STOP_WORDS
    | RU_STOP_WORDS
    | {
        "addlist",
        "joinchat",
        "invite",
        "авторский",
        "блог",
        "более",
        "больше",
        "будет",
        "было",
        "были",
        "быть",
        "ваш",
        "ваши",
        "ведет",
        "весь",
        "вместе",
        "внутри",
        "вопросы",
        "восемь",
        "врем",
        "время",
        "всего",
        "всех",
        "гибридный",
        "говорим",
        "году",
        "данные",
        "делаем",
        "делают",
        "день",
        "дизайн",
        "директор",
        "для",
        "друг",
        "других",
        "если",
        "есть",
        "ещё",
        "žеží",
        "žили",
        "žизни",
        "здесь",
        "знать",
        "иногда",
        "информация",
        "история",
        "итоги",
        "каждый",
        "канал",
        "канала",
        "канале",
        "каналу",
        "качестве",
        "когда",
        "компании",
        "компания",
        "контакт",
        "которая",
        "которое",
        "которые",
        "который",
        "кроме",
        "лучше",
        "люди",
        "можем",
        "может",
        "можно",
        "найти",
        "написать",
        "наш",
        "наша",
        "наше",
        "наши",
        "нашего",
        "немного",
        "нибудь",
        "новости",
        "нового",
        "обратной",
        "обучение",
        "около",
        "опыт",
        "очень",
        "пишем",
        "пишу",
        "подписаться",
        "подробнее",
        "пожалуйста",
        "пока",
        "полезн",
        "помощь",
        "после",
        "почта",
        "пресс",
        "просто",
        "работа",
        "работаем",
        "разбираем",
        "разное",
        "рассказываем",
        "реклама",
        "редакция",
        "сайт",
        "самое",
        "самые",
        "свои",
        "связи",
        "связь",
        "сделать",
        "сейчас",
        "сообщество",
        "сотрудничества",
        "сотрудничество",
        "ссылка",
        "ссылки",
        "статьи",
        "такие",
        "также",
        "телеграм",
        "тебе",
        "теперь",
        "только",
        "точка",
        "требования",
        "тут",
        "узнать",
        "хотите",
        "чата",
        "чате",
        "часть",
        "человек",
        "через",
        "что-то",
        "чтобы",
        "этом",
        "этот",
        "этого",
        "about",
        "author",
        "blog",
        "chat",
        "channel",
        "come",
        "contact",
        "from",
        "have",
        "here",
        "info",
        "more",
        "news",
        "only",
        "please",
        "read",
        "something",
        "team",
        "telegram",
        "that",
        "there",
        "they",
        "this",
        "what",
        "where",
        "which",
        "will",
        "with",
        "your",
        "https",
        "http",
        "com",
        "org",
        "html",
        "utm_source",
        "www",
    }
)

TOKEN = re.compile(r"[а-яёa-z]{4,}")

# A description is full of `t.me/+hrq31w2p1vuyogzi` invite hashes and
# similar, and a hash is distinctive by construction — it appears in one
# cluster because it appears exactly once anywhere, which is the highest
# score a rarity measure can give. Requiring a vowel is a crude test for
# "is this a word", and it costs nothing: the tokens it rejects were
# never going to name a crowd.
VOWELS = re.compile(r"[аеёиоуыэюяaeiouy]")

# Which alphabet a description is written in, which picks its
# lemmatizer. See `lemmatize`.
CYRILLIC_LETTER = re.compile(r"[а-яёА-ЯЁ]")
LATIN_LETTER = re.compile(r"[a-zA-Z]")

CHANNEL_COLUMNS = [
    "title",
    "username",
    "kind",
    "family_key",
    "cluster",
    "cluster_name",
    "stability",
    "partners",
    "link_weight",
    "inside_share",
    "fwd_inside",
    "mention_inside",
    "cluster_fwd",
    "cluster_mention",
    "depth",
    "dryness",
]

CLUSTER_COLUMNS = [
    "cluster",
    "name",
    "size",
    "internal_share",
    "fwd_internal",
    "mention_internal",
    "median_stability",
    "median_depth",
    "median_dryness",
    "top_kind",
    "members",
]

BRIDGE_COLUMNS = [
    "title",
    "username",
    "kind",
    "cluster",
    "cluster_name",
    "partners",
    "inside_share",
    "leans_to",
    "leans_share",
]

LABEL_NOTE = (
    "An arbitrary label with no meaning beyond this column — cluster 3 "
    "is not bigger, better or nearer to cluster 2 than to cluster 9, and "
    "the numbers change between runs. Only equality means anything: two "
    "channels with the same value were put in the same crowd."
)

CHANNEL_NOTES = {
    "family_key": (
        "The family of affiliated channels this one belongs to, as the "
        "smallest channel id in it. A label, not a main channel. "
        "References inside a family are dropped before clustering."
    ),
    "cluster": (
        "Which crowd this channel was placed in, over forwards and "
        "mentions together. " + LABEL_NOTE + " Blank means the channel "
        f"landed in a cluster under {MIN_CLUSTER} channels, has fewer "
        f"than {MIN_PARTNERS} distinct partners, or has no references at "
        "all in either direction."
    ),
    "cluster_name": (
        "The terms that are common in this cluster's titles and channel "
        "descriptions and rare in everyone else's. A handle for talking "
        "about the cluster, not a definition of it — read the members."
    ),
    "stability": (
        f"Over {STABILITY_RUNS} runs of Leiden from different random "
        "seeds, how often this channel landed with the same companions. "
        "1.00 means its placement is a fact about the graph; 0.50 means "
        "half the time it goes somewhere else, and its cluster is a coin "
        "toss the sheet is reporting as a result. The channels with the "
        "lowest values here are the interesting ones — they sit between "
        "crowds — but nothing about them should be quoted as membership."
    ),
    "partners": (
        "Distinct other seed channels this one references or is "
        "referenced by, in either direction, family excluded. This is "
        "the evidence the cluster label rests on."
    ),
    "link_weight": (
        "Total decayed, damped weight on this channel's edges — the "
        "number the clustering actually used. Read partners first; this "
        "is here to explain a surprise, not to be ranked on."
    ),
    "inside_share": (
        "Share of this channel's edge weight that stays inside its own "
        "cluster. Near 1 is a channel that only talks to its crowd; low "
        "is a connector. Blank without a cluster."
    ),
    "fwd_inside": (
        "The same share counting forwards only, against the same "
        "cluster. Blank when the channel has no forward edges."
    ),
    "mention_inside": (
        "The same share counting mentions only, against the same "
        "cluster. **Read against fwd_inside — the gap is the point of "
        "this sheet.** A channel that reposts inside its crowd but is "
        "mentioned from outside it is republished by its own and "
        "discussed by others, which a single clustering cannot say. "
        "Blank when the channel has no mention edges."
    ),
    "cluster_fwd": (
        "Which crowd the forward graph alone would have put this channel "
        "in. " + LABEL_NOTE + " Not comparable with the cluster column: "
        "these are labels from a separate partition of a smaller graph. "
        "Two channels sharing a value here repost the same crowd."
    ),
    "cluster_mention": (
        "Which crowd the mention graph alone would have put this channel "
        "in. " + LABEL_NOTE + " The two partitions agree far less than "
        "they look like they should; that disagreement is measured and "
        "printed when this script runs."
    ),
    "depth": (
        "How much this channel writes, from channel_style.xlsx — the "
        "mean of robust z-scores over post length, sentence length and "
        "how often it posts text at all. Repeated here so a cluster can "
        "be read for what it is like as well as who it talks to."
    ),
    "dryness": (
        "How technical this channel is, from channel_style.xlsx. Zero is "
        "the inventory median."
    ),
}

CLUSTER_NOTES = {
    "cluster": LABEL_NOTE,
    "name": (
        "Terms frequent in this cluster's titles and descriptions and "
        "rare elsewhere. A cluster whose name reads as nonsense is not a "
        "broken cluster — it may be a social circle rather than a "
        "subject, which is exactly the kind of thing this is for finding."
    ),
    "internal_share": (
        "Share of the edge weight touching this cluster that stays "
        "inside it. High means a closed crowd; low means the cluster "
        "exists but leaks."
    ),
    "fwd_internal": "The same, counting forwards only.",
    "mention_internal": (
        "The same, counting mentions only. A cluster that reposts itself "
        "and is mentioned from outside has a high fwd_internal and a low "
        "mention_internal."
    ),
    "median_stability": (
        "Median over members. A cluster under about 0.8 is not a stable "
        "finding — it is a region of the graph that Leiden cuts "
        "differently on different runs."
    ),
    "median_depth": (
        "Median depth of the members that have one. Blank when too few "
        "members were scorable in channel_style.xlsx."
    ),
    "median_dryness": "Median dryness of the members that have one.",
    "top_kind": (
        "The most common hand-reviewed kind among the members, and how "
        "many of them carry it."
    ),
    "members": (
        "The most connected members by edge weight — a handle for "
        "recognizing the cluster, not its full membership. The channels "
        "sheet has every row."
    ),
}

BRIDGE_NOTES = {
    "inside_share": (
        "Share of edge weight staying inside the channel's own cluster, "
        "ascending — the top of this sheet is the channels that connect "
        "crowds rather than sitting in one."
    ),
    "leans_to": (
        "The cluster taking the largest share of this channel's outside "
        "weight: who it connects its own crowd to."
    ),
    "leans_share": "That cluster's share of this channel's total weight.",
}

WIDTHS = {
    "title": 38,
    "username": 22,
    "kind": 12,
    "family_key": 13,
    "cluster_name": 34,
    "name": 36,
    "members": 70,
    "leans_to": 30,
    "top_kind": 16,
}

FORMATS = {
    "family_key": "0",
    "cluster": "0",
    "cluster_fwd": "0",
    "cluster_mention": "0",
    "stability": "0.00",
    "partners": "#,##0",
    "link_weight": "0.0",
    "inside_share": "0.00",
    "fwd_inside": "0.00",
    "mention_inside": "0.00",
    "depth": "0.00",
    "dryness": "0.00",
    "size": "#,##0",
    "internal_share": "0.00",
    "fwd_internal": "0.00",
    "mention_internal": "0.00",
    "median_stability": "0.00",
    "median_depth": "0.00",
    "median_dryness": "0.00",
    "leans_share": "0.00",
}


@dataclass
class Clustering:
    """Everything one run produced, for the map script to reuse."""

    graph: ig.Graph
    # Channel id -> cluster label, only for labels that survived the
    # MIN_CLUSTER and MIN_PARTNERS rules.
    cluster: dict[int, int]
    stability: dict[int, float]
    cluster_fwd: dict[int, int]
    cluster_mention: dict[int, int]
    names: dict[int, str]
    # Channel id -> (title, username, kind, about)
    meta: dict[int, tuple[str | None, str | None, str | None, str | None]]
    families: dict[int, int]
    style: pd.DataFrame
    # (a, b) -> weight per kind, a < b, family edges already dropped.
    pairs: dict[tuple[int, int], dict[str, float]] = field(
        default_factory=dict
    )
    dropped_family: int = 0
    kind_scale: float = 1.0


def decayed(published_at: datetime, now: datetime) -> float:
    days = (now - published_at).total_seconds() / 86400.0
    return math.pow(0.5, days / HALF_LIFE_DAYS)


def leiden(graph: ig.Graph, seed: int) -> list[int]:
    """One Leiden partition of a weighted graph, as a membership list."""
    partition = leidenalg.find_partition(
        graph,
        leidenalg.RBConfigurationVertexPartition,
        weights="weight",
        resolution_parameter=RESOLUTION,
        n_iterations=-1,
        seed=seed,
    )
    return list(partition.membership)


def subgraph(
    nodes: list[int],
    pairs: dict[tuple[int, int], dict[str, float]],
    kinds: tuple[str, ...],
    scale: dict[str, float],
) -> tuple[ig.Graph, list[int]]:
    """A weighted igraph over the named kinds, isolated nodes removed.

    Returns the graph and the channel ids of its vertices, in vertex
    order — igraph renumbers on deletion, so the mapping has to travel
    with the graph rather than be reconstructed from it.
    """
    index = {node: i for i, node in enumerate(nodes)}
    edges: list[tuple[int, int]] = []
    weights: list[float] = []
    for (a, b), by_kind in pairs.items():
        weight = sum(by_kind[kind] * scale[kind] for kind in kinds)
        if weight <= 0:
            continue
        edges.append((index[a], index[b]))
        weights.append(math.sqrt(weight))

    graph = ig.Graph(n=len(nodes), edges=edges)
    graph.es["weight"] = weights
    graph.vs["channel"] = nodes
    graph.delete_vertices(
        [vertex.index for vertex in graph.vs if vertex.degree() == 0]
    )
    return graph, list(graph.vs["channel"])


def stability_of(graph: ig.Graph, membership: list[int]) -> list[float]:
    """How often each vertex keeps the company it was finally given.

    Co-assignment frequency over the repeat runs, averaged across the
    vertex's own cluster-mates. A vertex alone in its cluster has no
    companions to keep, and scores 1 by convention — it is reported
    against `partners`, which is what says whether that means anything.
    """
    together: defaultdict[int, Counter[int]] = defaultdict(Counter)
    for seed in range(STABILITY_RUNS):
        run = leiden(graph, seed)
        by_label: defaultdict[int, list[int]] = defaultdict(list)
        for vertex, label in enumerate(run):
            by_label[label].append(vertex)
        for members in by_label.values():
            for vertex in members:
                together[vertex].update(members)

    final: defaultdict[int, list[int]] = defaultdict(list)
    for vertex, label in enumerate(membership):
        final[label].append(vertex)

    scores = [1.0] * graph.vcount()
    for members in final.values():
        for vertex in members:
            companions = [other for other in members if other != vertex]
            if not companions:
                continue
            scores[vertex] = sum(
                together[vertex][other] for other in companions
            ) / (len(companions) * STABILITY_RUNS)
    return scores


@cache
def _pipelines() -> tuple[Language, Language]:
    """The two lemmatizers, loaded once.

    ``parser`` and ``ner`` are excluded: nothing here reads a dependency
    tree or an entity span, and they are most of the load time. The
    Russian lemmatizer is *not* a lookup table — it picks between
    homographs by part of speech, so ``morphologizer`` and
    ``attribute_ruler`` have to stay or the lemmas come back wrong in
    exactly the ambiguous cases worth getting right.
    """
    return (
        spacy.load("ru_core_news_sm", exclude=["parser", "ner"]),
        spacy.load("en_core_web_sm", exclude=["parser", "ner"]),
    )


def lemmatize(texts: list[str]) -> list[list[str]]:
    """Each text's lemmas, lowercased and filtered, in the same order.

    Routed by alphabet rather than run through one model: the Russian
    pipeline reduces English to noise and the English one leaves Russian
    inflected, which is the whole problem this function exists to fix.
    A description that mixes the two — most of them do — goes to the
    model matching its majority, and the minority tokens come back
    close enough to unchanged to still count.

    Filtering happens on the lemma *and* the surface form. A stopword
    list written in dictionary forms would otherwise miss every inflected
    spelling, which is the same bug from the other side.
    """
    russian, english = _pipelines()
    groups: dict[Language, list[int]] = {russian: [], english: []}
    for index, text in enumerate(texts):
        cyrillic = len(CYRILLIC_LETTER.findall(text))
        latin = len(LATIN_LETTER.findall(text))
        groups[russian if cyrillic >= latin else english].append(index)

    out: list[list[str]] = [[] for _ in texts]
    for pipeline, indexes in groups.items():
        docs = pipeline.pipe(
            [texts[index] for index in indexes], batch_size=64
        )
        for index, doc in zip(indexes, docs, strict=True):
            out[index] = [
                lemma
                for token in doc
                if (lemma := token.lemma_.lower())
                and TOKEN.fullmatch(lemma)
                and VOWELS.search(lemma)
                and lemma not in STOPWORDS
                and token.text.lower() not in STOPWORDS
            ]
    return out


def name_clusters(
    members: dict[int, list[int]],
    meta: dict[int, tuple[str | None, str | None, str | None, str | None]],
) -> dict[int, str]:
    """Terms frequent in a cluster's own words and rare in everyone's.

    A log-odds against the rest of the corpus rather than a plain TF-IDF:
    with fifteen clusters, an inverse *document* frequency can only take
    fifteen values and stops separating anything.

    Counted over lemmas, so ``вакансия``, ``вакансии`` and ``вакансий``
    are one term with the weight of three rather than three terms each
    too thin to clear ``MIN_TERM_COUNT``. That threshold is what makes
    lemmatizing worth its two model loads: without it the inflections
    split a cluster's most characteristic word into forms that
    individually look like noise.
    """
    labels = sorted(members)
    documents = [
        " ".join(
            f"{meta.get(channel, (None,) * 4)[0] or ''} "
            f"{meta.get(channel, (None,) * 4)[3] or ''}"
            for channel in members[label]
        )
        for label in labels
    ]
    counts: dict[int, Counter[str]] = {
        label: Counter(terms)
        for label, terms in zip(labels, lemmatize(documents), strict=True)
    }

    total: Counter[str] = Counter()
    for bag in counts.values():
        total.update(bag)
    total_size = sum(total.values())

    names: dict[int, str] = {}
    for label, bag in counts.items():
        size = sum(bag.values())
        # Scaled to the cluster: three mentions is a reasonable floor
        # across fifty channels and an impossible one across eight, where
        # it left the smallest cluster with no name at all. Never below
        # two — a term used once is not a description of anything.
        floor = min(MIN_TERM_COUNT, max(2, len(members[label]) // 4))
        scored: list[tuple[float, str]] = []
        for term, count in bag.items():
            if count < floor:
                continue
            elsewhere = total[term] - count
            rest = total_size - size
            odds = (count / size) / ((elsewhere + 1) / (rest + 1))
            scored.append((odds, term))
        scored.sort(reverse=True)
        names[label] = ", ".join(term for _, term in scored[:NAME_TERMS])
    return names


def build(conn: psycopg.Connection[Any]) -> Clustering:
    """Load, weight, cluster, and measure how much of it to believe."""
    now = datetime.now(UTC)

    families = {channel_id: key for channel_id, key in conn.execute(FAMILIES)}
    meta = {
        row[0]: (row[1], row[2], row[3], row[4]) for row in conn.execute(NODES)
    }
    style = style_frame(conn)

    pairs: dict[tuple[int, int], dict[str, float]] = {}
    mass: Counter[str] = Counter()
    dropped_family = 0
    for src, dst, kind, published_at in conn.execute(EDGES):
        if families.get(src, src) == families.get(dst, dst):
            dropped_family += 1
            continue
        weight = decayed(published_at, now)
        key = (min(src, dst), max(src, dst))
        pairs.setdefault(key, dict.fromkeys(KINDS, 0.0))[kind] += weight
        mass[kind] += weight

    # Both kinds contribute the same total mass; see the docstring.
    scale = {
        "forward": 1.0,
        "mention": (mass["forward"] / mass["mention"])
        if mass["mention"]
        else 1.0,
    }

    nodes = sorted(meta)
    graph, channels = subgraph(nodes, pairs, KINDS, scale)
    membership = leiden(graph, seed=0)
    stability = stability_of(graph, membership)

    sizes = Counter(membership)
    partners = {channels[vertex.index]: vertex.degree() for vertex in graph.vs}
    cluster = {
        channels[vertex]: label
        for vertex, label in enumerate(membership)
        if sizes[label] >= MIN_CLUSTER
        and partners[channels[vertex]] >= MIN_PARTNERS
    }

    members: defaultdict[int, list[int]] = defaultdict(list)
    for channel, label in cluster.items():
        members[label].append(channel)

    def alone(kind: str) -> dict[int, int]:
        one, ids = subgraph(nodes, pairs, (kind,), scale)
        return {
            ids[vertex]: label
            for vertex, label in enumerate(leiden(one, seed=0))
        }

    return Clustering(
        graph=graph,
        cluster=cluster,
        stability={
            channels[vertex]: score for vertex, score in enumerate(stability)
        },
        cluster_fwd=alone("forward"),
        cluster_mention=alone("mention"),
        names=name_clusters(members, meta),
        meta=meta,
        families=families,
        style=style,
        pairs=pairs,
        dropped_family=dropped_family,
        kind_scale=scale["mention"],
    )


def shares(
    result: Clustering,
) -> tuple[dict[int, dict[str, float]], dict[int, dict[str, float]]]:
    """Per channel and per cluster: how much weight stays inside.

    Both are computed against the *combined* partition, including the
    per-kind splits. Measuring forwards against a forward-only partition
    and mentions against a mention-only one would compare each kind with
    a boundary drawn by itself, and every cluster would look closed.
    """
    inside: defaultdict[int, Counter[str]] = defaultdict(Counter)
    total: defaultdict[int, Counter[str]] = defaultdict(Counter)
    cluster_inside: defaultdict[int, Counter[str]] = defaultdict(Counter)
    cluster_total: defaultdict[int, Counter[str]] = defaultdict(Counter)

    for (a, b), by_kind in result.pairs.items():
        label_a = result.cluster.get(a)
        label_b = result.cluster.get(b)
        same = label_a is not None and label_a == label_b
        for kind in ("forward", "mention"):
            raw = by_kind[kind]
            if raw <= 0:
                continue
            weight = math.sqrt(raw)
            for node in (a, b):
                total[node][kind] += weight
                total[node]["all"] += weight
                if same:
                    inside[node][kind] += weight
                    inside[node]["all"] += weight
            for label in {label_a, label_b} - {None}:
                assert label is not None
                cluster_total[label][kind] += weight
                cluster_total[label]["all"] += weight
                if same:
                    cluster_inside[label][kind] += weight
                    cluster_inside[label]["all"] += weight

    def ratio(
        top: defaultdict[int, Counter[str]],
        bottom: defaultdict[int, Counter[str]],
    ) -> dict[int, dict[str, float]]:
        return {
            key: {
                kind: top[key][kind] / bottom[key][kind]
                for kind in ("all", "forward", "mention")
                if bottom[key][kind]
            }
            for key in bottom
        }

    return ratio(inside, total), ratio(cluster_inside, cluster_total)


def nmi(left: dict[int, int], right: dict[int, int]) -> float:
    """Normalized mutual information over the channels both partitions hold."""
    common = sorted(set(left) & set(right))
    if not common:
        return float("nan")
    a = [left[channel] for channel in common]
    b = [right[channel] for channel in common]
    n = len(common)
    ca, cb, cab = Counter(a), Counter(b), Counter(zip(a, b))
    ha = -sum(v / n * math.log(v / n) for v in ca.values())
    hb = -sum(v / n * math.log(v / n) for v in cb.values())
    mutual = sum(
        v / n * math.log((v / n) / ((ca[x] / n) * (cb[y] / n)))
        for (x, y), v in cab.items()
    )
    return mutual / ((ha + hb) / 2) if (ha + hb) else float("nan")


def main() -> None:
    with psycopg.connect(DSN) as conn:
        result = build(conn)

    node_share, cluster_share = shares(result)
    partners = {
        result.graph.vs[vertex.index]["channel"]: vertex.degree()
        for vertex in result.graph.vs
    }
    weight = {
        vertex["channel"]: sum(
            result.graph.es[edge]["weight"]
            for edge in result.graph.incident(vertex.index)
        )
        for vertex in result.graph.vs
    }

    outside_to = outward(result, node_share)

    channels = channel_sheet(result, node_share, partners, weight)
    clusters = cluster_sheet(result, cluster_share, weight)
    bridges = bridge_sheet(result, channels, outside_to)

    write({"channels": channels, "clusters": clusters, "bridges": bridges})
    write_gexf(result, node_share, partners)

    labelled = len(result.cluster)
    print(f"{result.graph.vcount()} channels with references -> {OUT}")
    print(
        f"{result.dropped_family} intra-family edges dropped; "
        f"mention weight scaled by {result.kind_scale:.3f}"
    )
    print(
        f"{len(clusters)} clusters of {MIN_CLUSTER}+ channels, "
        f"{labelled} channels labelled, "
        f"{len(result.meta) - labelled} left blank"
    )
    print(
        "forward-only vs mention-only partitions agree at NMI "
        f"{nmi(result.cluster_fwd, result.cluster_mention):.3f} "
        "— the gap is what fwd_inside and mention_inside are for"
    )
    shaky = sum(1 for value in result.stability.values() if value < 0.8)
    print(f"{shaky} channels under 0.80 stability — read those labels twice")
    print(f"graph -> {GEXF}")


def outward(
    result: Clustering, node_share: dict[int, dict[str, float]]
) -> dict[int, tuple[int | None, float]]:
    """For each channel, which other cluster takes most of its outside weight."""
    leaning: defaultdict[int, Counter[int]] = defaultdict(Counter)
    for (a, b), by_kind in result.pairs.items():
        label_a = result.cluster.get(a)
        label_b = result.cluster.get(b)
        if label_a == label_b:
            continue
        weight = math.sqrt(sum(by_kind.values()))
        if label_b is not None:
            leaning[a][label_b] += weight
        if label_a is not None:
            leaning[b][label_a] += weight

    out: dict[int, tuple[int | None, float]] = {}
    for channel, counter in leaning.items():
        if not counter:
            continue
        label, amount = counter.most_common(1)[0]
        total = sum(counter.values()) / max(
            1e-9, 1.0 - node_share.get(channel, {}).get("all", 0.0)
        )
        out[channel] = (label, amount / total if total else 0.0)
    return out


def channel_sheet(
    result: Clustering,
    node_share: dict[int, dict[str, float]],
    partners: dict[int, int],
    weight: dict[int, float],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    index: list[int] = []
    for channel, (title, username, kind, _) in sorted(result.meta.items()):
        label = result.cluster.get(channel)
        share = node_share.get(channel, {})
        style = (
            result.style.loc[channel]
            if channel in result.style.index
            else None
        )
        rows.append(
            {
                "title": title,
                "username": username,
                "kind": kind,
                "family_key": result.families.get(channel, channel),
                "cluster": label,
                "cluster_name": result.names.get(label)
                if label is not None
                else None,
                "stability": result.stability.get(channel),
                "partners": partners.get(channel, 0),
                "link_weight": weight.get(channel),
                "inside_share": share.get("all")
                if label is not None
                else None,
                "fwd_inside": share.get("forward")
                if label is not None
                else None,
                "mention_inside": (
                    share.get("mention") if label is not None else None
                ),
                "cluster_fwd": result.cluster_fwd.get(channel),
                "cluster_mention": result.cluster_mention.get(channel),
                "depth": None if style is None else style["depth"],
                "dryness": None if style is None else style["dryness"],
            }
        )
        index.append(channel)
    # Indexed by channel id: the bridges sheet needs to look a row back up
    # against the graph, and no visible column carries the id — a sheet
    # naming channel ids is a sheet that cannot be shown to anyone.
    return pd.DataFrame(rows, columns=CHANNEL_COLUMNS, index=pd.Index(index))


def cluster_sheet(
    result: Clustering,
    cluster_share: dict[int, dict[str, float]],
    weight: dict[int, float],
) -> pd.DataFrame:
    members: defaultdict[int, list[int]] = defaultdict(list)
    for channel, label in result.cluster.items():
        members[label].append(channel)

    rows: list[dict[str, Any]] = []
    for label, channels in members.items():
        share = cluster_share.get(label, {})
        top = sorted(channels, key=lambda c: -weight.get(c, 0.0))[:6]
        kinds = Counter(
            result.meta[c][2] for c in channels if result.meta[c][2]
        )
        scored = result.style.reindex(channels)
        rows.append(
            {
                "cluster": label,
                "name": result.names.get(label),
                "size": len(channels),
                "internal_share": share.get("all"),
                "fwd_internal": share.get("forward"),
                "mention_internal": share.get("mention"),
                "median_stability": pd.Series(
                    [result.stability[c] for c in channels]
                ).median(),
                "median_depth": scored["depth"].median(),
                "median_dryness": scored["dryness"].median(),
                "top_kind": (
                    f"{kinds.most_common(1)[0][0]} "
                    f"({kinds.most_common(1)[0][1]}/{len(channels)})"
                )
                if kinds
                else None,
                "members": " · ".join(
                    str(result.meta[c][0] or result.meta[c][1] or c)
                    for c in top
                ),
            }
        )
    frame = pd.DataFrame(rows, columns=CLUSTER_COLUMNS)
    return frame.sort_values("size", ascending=False, ignore_index=True)


def bridge_sheet(
    result: Clustering,
    channels: pd.DataFrame,
    outside_to: dict[int, tuple[int | None, float]],
) -> pd.DataFrame:
    """The channels that connect crowds, least inward-facing first."""
    wide = channels[
        channels["cluster"].notna()
        & (channels["partners"] >= BRIDGE_MIN_PARTNERS)
    ].copy()

    leans = [outside_to.get(channel) for channel in wide.index]
    wide["leans_to"] = [
        result.names.get(lean[0], "") if lean and lean[0] is not None else None
        for lean in leans
    ]
    wide["leans_share"] = [lean[1] if lean else None for lean in leans]

    frame = wide.sort_values("inside_share")
    return frame.reindex(columns=BRIDGE_COLUMNS).reset_index(drop=True)


def write(sheets: dict[str, pd.DataFrame]) -> None:
    """The workbook, unsorted where a sort would be an opinion."""
    OUT.parent.mkdir(parents=True, exist_ok=True)
    notes = {
        "channels": CHANNEL_NOTES,
        "clusters": CLUSTER_NOTES,
        "bridges": {**CHANNEL_NOTES, **BRIDGE_NOTES},
    }

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
            sheet = writer.sheets[name]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

            for index, column in enumerate(frame.columns, start=1):
                letter = get_column_letter(index)
                sheet.column_dimensions[letter].width = WIDTHS.get(column, 15)

                note = notes[name].get(column)
                if note:
                    sheet.cell(row=1, column=index).comment = Comment(
                        note, "clusters"
                    )

                number_format = FORMATS.get(column)
                if number_format:
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(
                            row=row, column=index
                        ).number_format = number_format


def write_gexf(
    result: Clustering,
    node_share: dict[int, dict[str, float]],
    partners: dict[int, int],
) -> None:
    """The same graph for Gephi, carrying the cluster it was given.

    Only channels and their references — no ids from the raw layer, no
    text. What `edges` holds is what this exports, which is what makes it
    safe to open in a tool that was not written for this project.
    """
    graph = nx.Graph()
    for vertex in result.graph.vs:
        channel = vertex["channel"]
        title, username, kind, _ = result.meta[channel]
        label = result.cluster.get(channel)
        style = (
            result.style.loc[channel]
            if channel in result.style.index
            else None
        )
        graph.add_node(
            str(channel),
            label=str(title or username or channel),
            username=username or "",
            kind=kind or "",
            cluster=str(label) if label is not None else "",
            cluster_name=result.names.get(label, "")
            if label is not None
            else "",
            stability=float(result.stability.get(channel, 0.0)),
            partners=int(partners.get(channel, 0)),
            inside_share=float(node_share.get(channel, {}).get("all", 0.0)),
            depth=float(depth_of(style, "depth")),
            dryness=float(depth_of(style, "dryness")),
        )

    for edge in result.graph.es:
        source = result.graph.vs[edge.source]["channel"]
        target = result.graph.vs[edge.target]["channel"]
        graph.add_edge(str(source), str(target), weight=float(edge["weight"]))

    GEXF.parent.mkdir(parents=True, exist_ok=True)
    nx.write_gexf(graph, GEXF)


def depth_of(style: pd.Series[Any] | None, column: str) -> float:
    """A style score as a number GEXF can hold; 0 where it is unmeasured."""
    if style is None:
        return 0.0
    value = style[column]
    return 0.0 if pd.isna(value) else float(value)


if __name__ == "__main__":
    main()
