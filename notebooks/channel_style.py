"""How each seed channel writes: length, density, code, emoji — as xlsx.

Exploratory tooling: not part of the package, not spec'd, not tested. Run
it, sort the sheet, change the constants, run it again.

    uv sync --group data
    uv run notebooks/channel_style.py

This is the "depth and dryness" axis, and it deliberately uses no model
of any kind. Median post length runs from 137 characters at the 5th
percentile to 1913 at the 95th across this inventory — a 14x spread that
separates the essayists from the meme feeds far more sharply than any
embedding would, and it costs one pass over text already collected.

**Forwarded posts are excluded from every style column.** A repost
carries the text of whoever wrote it, so measuring it describes the
channel that was copied, not the one doing the copying — and an
aggregator, whose whole output is other people's posts, would come out
looking exactly like the crowd it aggregates. What a repost says about
this channel is that it happened, which is what ``fwd_share`` is for; the
words in it belong to someone else and are dropped.

Everything else is a median or a rate per post, so the differing depth of
the backfill does not bias it — the same argument the scorecard makes.
``posts`` is a collection depth as much as a publishing rate, and is the
one column that must not be compared between channels walked to
different cutoffs.

The two scores at the end are the point of the sheet, and they are
deliberately two rather than one. Depth is how much a channel writes;
dryness is how technical it is about it. They are not the same axis and a
channel can be high on either alone — a long conversational personal
channel is deep and wet, a terse changelog is dry and shallow. Each is
the mean of a handful of robust z-scores over columns that are all in the
sheet, so a surprising score can always be taken apart into the numbers
that produced it. Deliberately not a principal component: the sign of a
component flips between runs, and no reader can be told what it means.
"""

import math
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import psycopg
from _db import dsn
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

DSN = dsn()
OUT = Path(__file__).resolve().parent.parent / "data" / "channel_style.xlsx"
SHEET = "channels"

# Below this many of the channel's *own* text posts, the style columns
# and both scores are left blank rather than reported. A median over a
# handful of posts is not a writing style, and a number on the screen
# gets sorted on whatever it means. At 50, 406 of 544 channels qualify.
MIN_TEXT_POSTS = 50

# Characters that end a sentence, and the letter classes that decide how
# much of a channel's prose is jargon. Written as literal ranges rather
# than \p{...} classes, which Postgres regex does not have.
#
# A newline ends a sentence here as surely as a full stop does, and
# leaving it out was measurably wrong rather than merely imprecise: a
# vacancy listing is a bulleted list with no full stops in it, so it came
# out as one 1300-character sentence and job feeds ranked as the deepest
# writing in the inventory. Telegram prose breaks lines where prose
# elsewhere would punctuate.
SENTENCE_END = "[.!?…\n]+"
LATIN = "[A-Za-z]"
CYRILLIC = "[А-Яа-яЁё]"
EMPHATIC = "[!?]"
# Emoji, approximately: the pictographic planes plus the older symbol
# block that carries ✅ ⚡ ☝ and friends. Approximate on purpose — the
# exact set is a Unicode property table, and this column only has to
# separate a channel that uses emoji from one that does not.
EMOJI = "[\U0001f300-\U0001faff☀-➿]"

CHANNELS = """
    SELECT c.tg_id, c.title, c.username, c.kind::text
    FROM channels c
    WHERE c.status = 'seed' AND NOT c.is_chat
"""

FAMILIES = """
    SELECT channel_id, family_key
    FROM channel_families
"""

# What the channel published, all posts, forwards included. Service
# messages are excluded: "channel photo changed" is not a post.
VOLUME = """
    SELECT
        r.channel_id,
        COUNT(*) AS posts,
        AVG((jsonb_typeof(r.payload->'fwd_from') = 'object')::int)::float8
            AS fwd_share,
        AVG((COALESCE(r.payload->>'message', '') <> '')::int)::float8
            AS text_share
    FROM raw_messages r
    JOIN channels c ON c.tg_id = r.channel_id
                   AND c.status = 'seed' AND NOT c.is_chat
    WHERE r.payload->>'_' = 'Message'
    GROUP BY r.channel_id
"""

# The style pass, over the channel's own text posts only.
#
# Entity offsets and lengths are in UTF-16 code units and `length()`
# counts characters, so `quote_share` drifts on text heavy in emoji or
# other astral-plane characters. It is a ratio of two large sums over
# prose, and the drift is well under the differences it is read for —
# recorded here rather than corrected, because correcting it means
# re-encoding every post to count units nobody else in this project uses.
#
# The `jsonb_typeof` guard on entities is not defensive noise: the
# payload stores an absent field as JSON `null`, and
# `jsonb_array_elements` raises on a scalar rather than returning nothing.
STYLE = """
    WITH own AS (
        SELECT
            r.channel_id,
            r.payload->>'message' AS txt,
            length(r.payload->>'message') AS chars,
            CASE WHEN jsonb_typeof(r.payload->'entities') = 'array'
                 THEN r.payload->'entities' ELSE '[]'::jsonb END AS ents
        FROM raw_messages r
        JOIN channels c ON c.tg_id = r.channel_id
                       AND c.status = 'seed' AND NOT c.is_chat
        WHERE r.payload->>'_' = 'Message'
          AND jsonb_typeof(r.payload->'fwd_from') <> 'object'
          AND COALESCE(r.payload->>'message', '') <> ''
    ), counted AS (
        SELECT
            channel_id,
            chars,
            GREATEST(regexp_count(txt, %(sentence_end)s), 1) AS sentences,
            regexp_count(txt, %(latin)s) AS latin,
            regexp_count(txt, %(cyrillic)s) AS cyrillic,
            regexp_count(txt, %(emphatic)s) AS emphatic,
            regexp_count(txt, %(emoji)s) AS emoji,
            (SELECT COUNT(*) FROM jsonb_array_elements(ents) e
              WHERE e->>'_' IN ('MessageEntityCode', 'MessageEntityPre'))
                AS n_code,
            (SELECT COUNT(*) FROM jsonb_array_elements(ents) e
              WHERE e->>'_' IN ('MessageEntityUrl', 'MessageEntityTextUrl'))
                AS n_link,
            (SELECT COUNT(*) FROM jsonb_array_elements(ents) e
              WHERE e->>'_' = 'MessageEntityCustomEmoji') AS n_custom,
            (SELECT COALESCE(SUM((e->>'length')::int), 0)
               FROM jsonb_array_elements(ents) e
              WHERE e->>'_' = 'MessageEntityBlockquote') AS quoted
        FROM own
    )
    SELECT
        channel_id,
        COUNT(*) AS text_posts,
        percentile_cont(0.5) WITHIN GROUP (ORDER BY chars) AS median_len,
        percentile_cont(0.9) WITHIN GROUP (ORDER BY chars) AS p90_len,
        percentile_cont(0.5) WITHIN GROUP (
            ORDER BY chars::float8 / sentences
        ) AS sentence_len,
        LEAST(SUM(quoted)::float8 / NULLIF(SUM(chars), 0), 1.0)
            AS quote_share,
        AVG((n_code > 0)::int)::float8 AS code_share,
        SUM(latin)::float8 / NULLIF(SUM(latin + cyrillic), 0)
            AS latin_share,
        1000.0 * SUM(n_link) / NULLIF(SUM(chars), 0) AS link_density,
        1000.0 * SUM(n_custom + emoji) / NULLIF(SUM(chars), 0)
            AS emoji_density,
        1000.0 * SUM(emphatic) / NULLIF(SUM(chars), 0) AS emphatic_density
    FROM counted
    GROUP BY channel_id
"""

COLUMNS = [
    "title",
    "username",
    "kind",
    "family_key",
    "posts",
    "text_posts",
    "text_share",
    "fwd_share",
    "median_len",
    "p90_len",
    "sentence_len",
    "quote_share",
    "code_share",
    "latin_share",
    "link_density",
    "emoji_density",
    "emphatic_density",
    "depth",
    "dryness",
]

# What each score averages, and which way each input points. The
# transform is what makes a robust z meaningful on a heavy-tailed count:
# lengths and densities are log-shaped, a share is not.
DEPTH_INPUTS = (
    ("median_len", "log", +1),
    ("sentence_len", "log", +1),
    ("text_share", "none", +1),
)
DRYNESS_INPUTS = (
    ("code_share", "sqrt", +1),
    ("latin_share", "none", +1),
    ("emoji_density", "log", -1),
    ("emphatic_density", "log", -1),
)

# A robust z is divided by the MAD, and a column whose MAD is zero — more
# than half the channels sharing one value — would divide by nothing.
# Such a column contributes nothing to the score rather than infinity.
MAD_TO_SIGMA = 1.4826

# One channel with a hundred code blocks should not decide the mean of
# four z-scores, so each is clipped before averaging.
Z_CLIP = 4.0

# Above this share of latin letters a channel is probably not writing in
# Russian at all, and `latin_share` is measuring its language rather than
# its jargon. Reported at the end rather than corrected for.
LATIN_IS_LANGUAGE = 0.7

STYLE_NOTE = (
    "Computed over this channel's own text posts only — forwards are "
    "excluded, because a repost carries the text of whoever wrote it. "
    f"Blank under {MIN_TEXT_POSTS} own text posts."
)

NOTES = {
    "family_key": (
        "The family of affiliated channels this one belongs to, as the "
        "smallest channel id in it. A label, not a main channel. A "
        "channel with no confirmed affiliation is its own family, so the "
        "key is its own id."
    ),
    "posts": (
        "Posts in the collected history, service messages excluded, "
        "forwards included. How deep that history goes is the backfill's "
        "cutoff and can differ per channel, so this is a collection "
        "depth as much as a publishing rate — the one column here that "
        "must not be compared between channels walked to different "
        "cutoffs."
    ),
    "text_posts": (
        "The channel's own posts that carry text: the denominator of "
        "every style column to the right of here. A media channel can "
        "have many posts and few of these."
    ),
    "text_share": (
        "Share of all posts that carry any text at all. Low means a "
        "channel of images, videos or files with captions at most."
    ),
    "fwd_share": (
        "Share of all posts that are forwards. This is what a repost "
        "says about the channel doing it; the words in the repost belong "
        "to someone else and are excluded from every other column."
    ),
    "median_len": "Median length in characters of an own text post. "
    + STYLE_NOTE,
    "p90_len": (
        "90th percentile length. Read against median_len: a channel with "
        "a 300-character median and a 3000-character p90 posts notes and "
        "occasional essays, which is a different habit from one that "
        "always writes 900. " + STYLE_NOTE
    ),
    "sentence_len": (
        "Median over posts of (characters / sentences), splitting on "
        ".!?… — how long a sentence this channel writes. " + STYLE_NOTE
    ),
    "quote_share": (
        "Share of characters sitting inside a blockquote: how much of "
        "the channel is quoting rather than writing. Approximate — "
        "Telegram counts entity offsets in UTF-16 units and this counts "
        "characters, which drifts on emoji-heavy text. " + STYLE_NOTE
    ),
    "code_share": (
        "Share of own text posts carrying a code span or block, as "
        "Telegram itself marked it. " + STYLE_NOTE
    ),
    "latin_share": (
        "Latin letters as a share of all letters — a proxy for technical "
        "jargon in Russian prose. CAVEAT: for a channel that writes in "
        "English this measures its language, not its jargon, and its "
        "dryness score is not comparable. " + STYLE_NOTE
    ),
    "link_density": "Links per 1000 characters. " + STYLE_NOTE,
    "emoji_density": (
        "Emoji per 1000 characters, custom and unicode together. The "
        "unicode set is approximate — it separates channels that use "
        "emoji from ones that do not, and is not an exact count. " + STYLE_NOTE
    ),
    "emphatic_density": (
        "Exclamation and question marks per 1000 characters. " + STYLE_NOTE
    ),
    "depth": (
        "How much this channel writes, as the mean of robust z-scores "
        "over log(median_len), log(sentence_len) and text_share — all "
        "three are columns in this sheet, so any score can be taken "
        "apart into what produced it. Zero is the inventory median, and "
        "the unit is roughly one standard deviation. Independent of "
        "dryness: a long chatty personal channel scores high here and "
        "low there. " + STYLE_NOTE
    ),
    "dryness": (
        "How technical this channel is, as the mean of robust z-scores "
        "over sqrt(code_share), latin_share, minus log(emoji_density) "
        "and minus log(emphatic_density). Zero is the inventory median. "
        "Read together with latin_share's caveat for channels not "
        "written in Russian. " + STYLE_NOTE
    ),
}

WIDTHS = {"title": 38, "username": 22, "kind": 12, "family_key": 13}

INT_FORMAT = "#,##0"
FORMATS = {
    "family_key": "0",
    "posts": INT_FORMAT,
    "text_posts": INT_FORMAT,
    "text_share": "0.00",
    "fwd_share": "0.00",
    "median_len": INT_FORMAT,
    "p90_len": INT_FORMAT,
    "sentence_len": "0.0",
    "quote_share": "0.000",
    "code_share": "0.000",
    "latin_share": "0.000",
    "link_density": "0.00",
    "emoji_density": "0.00",
    "emphatic_density": "0.00",
    "depth": "0.00",
    "dryness": "0.00",
}


def robust_z(values: pd.Series, transform: str) -> pd.Series:
    """Median-centred, MAD-scaled, clipped — never mean and stdev.

    The same reason every baseline in this project is a median: the
    channels this is meant to find are the tails, and a mean moved by
    them measures against a ruler they bent themselves.
    """
    if transform == "log":
        scaled = np.log1p(values.astype(float))
    elif transform == "sqrt":
        scaled = np.sqrt(values.astype(float))
    else:
        scaled = values.astype(float)

    median = scaled.median()
    mad = (scaled - median).abs().median()
    if not mad or math.isnan(mad):
        return pd.Series(np.nan, index=values.index)
    return ((scaled - median) / (MAD_TO_SIGMA * mad)).clip(-Z_CLIP, Z_CLIP)


def score(
    frame: pd.DataFrame, inputs: tuple[tuple[str, str, int], ...]
) -> pd.Series:
    """The mean of the signed robust z-scores of the named columns."""
    parts = [
        sign * robust_z(frame[column], transform)
        for column, transform, sign in inputs
    ]
    return pd.concat(parts, axis=1).mean(axis=1)


def style_frame(conn: psycopg.Connection[Any]) -> pd.DataFrame:
    """Every style column and both scores, indexed by channel id.

    Public because ``clusters.py`` reads it: a cluster's median depth and
    dryness are what turn a set of channel ids into something a person
    can describe, and recomputing this query there would be the same
    measurement written twice, free to drift apart.
    """
    params = {
        "sentence_end": SENTENCE_END,
        "latin": LATIN,
        "cyrillic": CYRILLIC,
        "emphatic": EMPHATIC,
        "emoji": EMOJI,
    }

    families = {channel_id: key for channel_id, key in conn.execute(FAMILIES)}
    channels = conn.execute(CHANNELS).fetchall()
    volume = {row[0]: row[1:] for row in conn.execute(VOLUME)}
    style = {row[0]: row[1:] for row in conn.execute(STYLE, params)}

    rows: list[dict[str, Any]] = []
    index: list[int] = []
    for tg_id, title, username, kind in channels:
        posts, fwd_share, text_share = volume.get(tg_id, (0, None, None))
        (
            text_posts,
            median_len,
            p90_len,
            sentence_len,
            quote_share,
            code_share,
            latin_share,
            link_density,
            emoji_density,
            emphatic_density,
        ) = style.get(tg_id, (0,) + (None,) * 9)

        thin = (text_posts or 0) < MIN_TEXT_POSTS
        blank: Any = None
        index.append(tg_id)
        rows.append(
            {
                "title": title,
                "username": username,
                "kind": kind,
                "family_key": families.get(tg_id, tg_id),
                "posts": posts,
                "text_posts": text_posts,
                "text_share": text_share,
                "fwd_share": fwd_share,
                "median_len": blank if thin else median_len,
                "p90_len": blank if thin else p90_len,
                "sentence_len": blank if thin else sentence_len,
                "quote_share": blank if thin else quote_share,
                "code_share": blank if thin else code_share,
                "latin_share": blank if thin else latin_share,
                "link_density": blank if thin else link_density,
                "emoji_density": blank if thin else emoji_density,
                "emphatic_density": blank if thin else emphatic_density,
            }
        )

    frame = pd.DataFrame(rows, columns=COLUMNS, index=pd.Index(index))

    # Scored over the channels that qualify, so the median a z-score is
    # taken against is the median of measured channels — not one dragged
    # towards zero by every channel too thin to measure.
    scored = frame["median_len"].notna()
    frame.loc[scored, "depth"] = score(frame[scored], DEPTH_INPUTS)
    frame.loc[scored, "dryness"] = score(frame[scored], DRYNESS_INPUTS)
    return frame


def main() -> None:
    with psycopg.connect(DSN) as conn:
        frame = style_frame(conn)

    scored = frame["median_len"].notna()
    write(frame)

    latin_heavy = int((frame["latin_share"] > LATIN_IS_LANGUAGE).sum())
    print(f"{len(frame)} seed channels -> {OUT}")
    print(
        f"{int(scored.sum())} scored, "
        f"{int((~scored).sum())} under {MIN_TEXT_POSTS} own text posts"
    )
    print(
        f"{latin_heavy} channels over {LATIN_IS_LANGUAGE:.0%} latin letters "
        "— probably not Russian, dryness not comparable for them"
    )


def write(frame: pd.DataFrame) -> None:
    """The sheet, unsorted: header frozen, autofilter on, notes attached.

    No sort order is baked in — the point of an xlsx is that the reader
    picks the column and the direction, and any default here would be an
    opinion about what matters.
    """
    OUT.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name=SHEET, index=False)
        sheet = writer.sheets[SHEET]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for index, column in enumerate(frame.columns, start=1):
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = WIDTHS.get(column, 15)

            note = NOTES.get(column)
            if note:
                header = sheet.cell(row=1, column=index)
                header.comment = Comment(note, "channel_style")

            number_format = FORMATS.get(column)
            if number_format:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(
                        row=row, column=index
                    ).number_format = number_format


if __name__ == "__main__":
    main()
